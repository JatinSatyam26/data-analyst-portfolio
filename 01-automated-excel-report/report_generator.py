"""
report_generator.py
────────────────────────────────────────────────────────────────
Automated Excel Report Generator
Author : Jatin Prasad
Purpose: Ingests raw CSV sales data, cleans it, runs analysis,
         and produces a fully-formatted multi-sheet Excel report
         — zero manual steps required.

Sheets produced:
  1. Executive Summary   – KPI cards + status breakdown
  2. Monthly Trends      – revenue & orders by month (+ bar chart)
  3. Product Analysis    – revenue & quantity by product (+ bar chart)
  4. Regional Analysis   – revenue breakdown by region  (+ bar chart)
  5. Sales Rep Leaderboard – top performers ranked
  6. Raw Data            – cleaned source data for audit trail
────────────────────────────────────────────────────────────────
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import datetime
import sys
import os

# ── Palette ────────────────────────────────────────────────────
DARK_TEAL   = "1A5276"   # header bg
MID_TEAL    = "148F77"   # section header
LIGHT_TEAL  = "D1F2EB"   # alternating row
ACCENT      = "F39C12"   # KPI value highlight
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F3F4"
DARK_GRAY   = "2C3E50"

# ── Helpers ────────────────────────────────────────────────────
def _border(style="thin"):
    s = Side(style=style, color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_font(size=11, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def _body_font(size=10, bold=False, color=DARK_GRAY):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center")

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def autofit_cols(ws, min_w=10, max_w=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)

def write_table_header(ws, row, cols, col_start=1):
    """Write a styled header row for a data table."""
    for i, col_name in enumerate(cols, start=col_start):
        cell = ws.cell(row=row, column=i, value=col_name)
        cell.font      = _header_font(size=10)
        cell.fill      = _fill(MID_TEAL)
        cell.alignment = _center()
        cell.border    = _border()

def write_data_row(ws, row_idx, values, col_start=1, shade=False):
    """Write a data row with alternating shading."""
    bg = LIGHT_TEAL if shade else WHITE
    for i, val in enumerate(values, start=col_start):
        cell = ws.cell(row=row_idx, column=i, value=val)
        cell.font      = _body_font()
        cell.fill      = _fill(bg)
        cell.alignment = _center()
        cell.border    = _border()

def section_header(ws, row, col, text, span=1, bg=DARK_TEAL):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + span - 1)
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = _header_font(size=12)
    cell.fill      = _fill(bg)
    cell.alignment = _center()

# ── Data Loading & Cleaning ────────────────────────────────────
def load_and_clean(filepath):
    print(f"📂  Loading data from: {filepath}")
    df = pd.read_csv(filepath)

    original_rows = len(df)

    # Parse dates
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")

    # Drop rows with null critical fields
    df.dropna(subset=["Date", "Revenue", "Product", "Region"], inplace=True)

    # Remove duplicated Order_IDs (keep first)
    df.drop_duplicates(subset=["Order_ID"], keep="first", inplace=True)

    # Remove negative revenue
    df = df[df["Revenue"] > 0]

    # Standardise text columns
    for col in ["Product", "Region", "Sales_Rep", "Order_Status"]:
        df[col] = df[col].str.strip().str.title()

    # Derived columns
    df["Month"]        = df["Date"].dt.to_period("M").dt.to_timestamp()
    df["Month_Label"]  = df["Date"].dt.strftime("%b %Y")
    df["Year"]         = df["Date"].dt.year

    # Only completed orders count as revenue
    df["Effective_Revenue"] = df.apply(
        lambda r: r["Revenue"] if r["Order_Status"] == "Completed" else 0, axis=1
    )

    cleaned_rows = len(df)
    removed = original_rows - cleaned_rows
    print(f"✅  Data cleaned: {cleaned_rows} rows retained, {removed} rows removed")
    return df

# ── Analysis ──────────────────────────────────────────────────
def run_analysis(df):
    completed = df[df["Order_Status"] == "Completed"]

    kpis = {
        "Total Orders":        len(df),
        "Completed Orders":    len(completed),
        "Total Revenue ($)":   round(df["Effective_Revenue"].sum(), 2),
        "Avg Order Value ($)": round(completed["Revenue"].mean(), 2) if len(completed) else 0,
        "Top Product":         completed.groupby("Product")["Revenue"].sum().idxmax() if len(completed) else "N/A",
        "Top Region":          completed.groupby("Region")["Revenue"].sum().idxmax() if len(completed) else "N/A",
        "Return Rate (%)":     round(len(df[df["Order_Status"] == "Returned"]) / len(df) * 100, 1),
    }

    monthly = (df.groupby("Month_Label", sort=False)
                 .agg(Orders=("Order_ID", "count"),
                      Revenue=("Effective_Revenue", "sum"))
                 .reset_index()
                 .rename(columns={"Month_Label": "Month"}))
    # Sort by actual month
    monthly["_sort"] = pd.to_datetime(monthly["Month"], format="%b %Y")
    monthly = monthly.sort_values("_sort").drop(columns="_sort")
    monthly["Revenue"] = monthly["Revenue"].round(2)

    by_product = (completed.groupby("Product")
                            .agg(Total_Revenue=("Revenue", "sum"),
                                 Total_Qty=("Quantity", "sum"),
                                 Avg_Price=("Unit_Price", "mean"))
                            .reset_index()
                            .sort_values("Total_Revenue", ascending=False))
    by_product["Total_Revenue"] = by_product["Total_Revenue"].round(2)
    by_product["Avg_Price"]     = by_product["Avg_Price"].round(2)

    by_region = (completed.groupby("Region")
                           .agg(Total_Revenue=("Revenue", "sum"),
                                Orders=("Order_ID", "count"))
                           .reset_index()
                           .sort_values("Total_Revenue", ascending=False))
    by_region["Total_Revenue"]  = by_region["Total_Revenue"].round(2)
    by_region["Revenue_Share"]  = (by_region["Total_Revenue"] /
                                   by_region["Total_Revenue"].sum() * 100).round(1)

    by_rep = (completed.groupby("Sales_Rep")
                        .agg(Total_Revenue=("Revenue", "sum"),
                             Orders=("Order_ID", "count"),
                             Avg_Order=("Revenue", "mean"))
                        .reset_index()
                        .sort_values("Total_Revenue", ascending=False)
                        .reset_index(drop=True))
    by_rep["Rank"]          = by_rep.index + 1
    by_rep["Total_Revenue"] = by_rep["Total_Revenue"].round(2)
    by_rep["Avg_Order"]     = by_rep["Avg_Order"].round(2)

    status_counts = df["Order_Status"].value_counts().reset_index()
    status_counts.columns = ["Status", "Count"]

    print("✅  Analysis complete")
    return kpis, monthly, by_product, by_region, by_rep, status_counts

# ── Sheet Builders ─────────────────────────────────────────────
def build_summary_sheet(ws, kpis, status_counts, report_date):
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 20

    # Title banner
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value     = "📊  ANNUAL SALES PERFORMANCE REPORT — 2024"
    cell.font      = Font(name="Calibri", size=16, bold=True, color=WHITE)
    cell.fill      = _fill(DARK_TEAL)
    cell.alignment = _center()

    ws.merge_cells("A2:H2")
    sub = ws["A2"]
    sub.value     = f"Generated automatically on {report_date}   |   Ira A. Fulton Schools of Engineering"
    sub.font      = Font(name="Calibri", size=9, italic=True, color="777777")
    sub.alignment = _center()

    # KPI cards — two rows of 4
    kpi_items = list(kpis.items())
    kpi_layout = [(4, 1), (4, 3), (4, 5), (4, 7),
                  (7, 1), (7, 3), (7, 5), (7, 7)]

    for (r, c), (label, value) in zip(kpi_layout, kpi_items):
        # Label row
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+1)
        lc = ws.cell(row=r, column=c, value=label)
        lc.font      = Font(name="Calibri", size=9, bold=True, color=WHITE)
        lc.fill      = _fill(MID_TEAL)
        lc.alignment = _center()
        ws.row_dimensions[r].height = 18

        # Value row
        ws.merge_cells(start_row=r+1, start_column=c, end_row=r+1, end_column=c+1)
        vc = ws.cell(row=r+1, column=c, value=value)
        vc.font      = Font(name="Calibri", size=14, bold=True, color=ACCENT)
        vc.fill      = _fill(LIGHT_GRAY)
        vc.alignment = _center()
        vc.border    = _border()
        ws.row_dimensions[r+1].height = 28

    # Order Status breakdown
    section_header(ws, 10, 1, "ORDER STATUS BREAKDOWN", span=3)
    write_table_header(ws, 11, ["Status", "Count", "Share (%)"], col_start=1)
    total_orders = status_counts["Count"].sum()
    for i, (_, row) in enumerate(status_counts.iterrows()):
        share = round(row["Count"] / total_orders * 100, 1)
        write_data_row(ws, 12 + i, [row["Status"], row["Count"], f"{share}%"], shade=i % 2 == 0)

    for c, w in zip(range(1, 9), [14, 14, 14, 14, 14, 14, 14, 14]):
        set_col_width(ws, c, w)

def build_monthly_sheet(ws, monthly):
    ws.sheet_view.showGridLines = False

    section_header(ws, 1, 1, "MONTHLY REVENUE & ORDER TRENDS", span=3, bg=DARK_TEAL)
    write_table_header(ws, 2, ["Month", "Orders", "Revenue ($)"])

    for i, (_, row) in enumerate(monthly.iterrows()):
        write_data_row(ws, 3 + i, [row["Month"], row["Orders"], row["Revenue"]], shade=i % 2 == 0)

    # Bar chart
    data_rows = len(monthly)
    chart = BarChart()
    chart.type    = "col"
    chart.title   = "Monthly Revenue (Completed Orders)"
    chart.y_axis.title = "Revenue ($)"
    chart.x_axis.title = "Month"
    chart.height  = 14
    chart.width   = 24
    chart.style   = 10

    data_ref   = Reference(ws, min_col=3, min_row=2, max_row=2 + data_rows)
    labels_ref = Reference(ws, min_col=1, min_row=3, max_row=2 + data_rows)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(labels_ref)
    ws.add_chart(chart, f"E2")

    autofit_cols(ws)

def build_product_sheet(ws, by_product):
    ws.sheet_view.showGridLines = False

    section_header(ws, 1, 1, "PRODUCT PERFORMANCE ANALYSIS", span=4, bg=DARK_TEAL)
    write_table_header(ws, 2, ["Product", "Total Revenue ($)", "Units Sold", "Avg Unit Price ($)"])

    for i, (_, row) in enumerate(by_product.iterrows()):
        write_data_row(ws, 3 + i,
                       [row["Product"], row["Total_Revenue"],
                        row["Total_Qty"], row["Avg_Price"]],
                       shade=i % 2 == 0)

    data_rows = len(by_product)
    chart = BarChart()
    chart.type   = "bar"   # horizontal
    chart.title  = "Revenue by Product"
    chart.y_axis.title = "Product"
    chart.x_axis.title = "Revenue ($)"
    chart.height = 14
    chart.width  = 22
    chart.style  = 10

    data_ref   = Reference(ws, min_col=2, min_row=2, max_row=2 + data_rows)
    labels_ref = Reference(ws, min_col=1, min_row=3, max_row=2 + data_rows)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(labels_ref)
    ws.add_chart(chart, "F2")

    autofit_cols(ws)

def build_regional_sheet(ws, by_region):
    ws.sheet_view.showGridLines = False

    section_header(ws, 1, 1, "REGIONAL PERFORMANCE ANALYSIS", span=4, bg=DARK_TEAL)
    write_table_header(ws, 2, ["Region", "Total Revenue ($)", "Orders", "Revenue Share (%)"])

    for i, (_, row) in enumerate(by_region.iterrows()):
        write_data_row(ws, 3 + i,
                       [row["Region"], row["Total_Revenue"],
                        row["Orders"], f"{row['Revenue_Share']}%"],
                       shade=i % 2 == 0)

    data_rows = len(by_region)
    chart = BarChart()
    chart.type   = "col"
    chart.title  = "Revenue by Region"
    chart.height = 12
    chart.width  = 20
    chart.style  = 10

    data_ref   = Reference(ws, min_col=2, min_row=2, max_row=2 + data_rows)
    labels_ref = Reference(ws, min_col=1, min_row=3, max_row=2 + data_rows)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(labels_ref)
    ws.add_chart(chart, "F2")

    autofit_cols(ws)

def build_leaderboard_sheet(ws, by_rep):
    ws.sheet_view.showGridLines = False

    section_header(ws, 1, 1, "SALES REP LEADERBOARD", span=5, bg=DARK_TEAL)
    write_table_header(ws, 2, ["Rank", "Sales Rep", "Total Revenue ($)", "Orders Closed", "Avg Order Value ($)"])

    medal = {1: "🥇", 2: "🥈", 3: "🥉"}
    for i, (_, row) in enumerate(by_rep.iterrows()):
        rank_label = f"{medal.get(int(row['Rank']), '')} #{int(row['Rank'])}"
        write_data_row(ws, 3 + i,
                       [rank_label, row["Sales_Rep"], row["Total_Revenue"],
                        row["Orders"], row["Avg_Order"]],
                       shade=i % 2 == 0)
        # Highlight top 3
        if row["Rank"] <= 3:
            for col in range(1, 6):
                ws.cell(row=3 + i, column=col).font = Font(
                    name="Calibri", size=10, bold=True, color=DARK_GRAY)

    autofit_cols(ws)

def build_raw_data_sheet(ws, df):
    ws.sheet_view.showGridLines = True

    display_cols = ["Order_ID", "Date", "Product", "Region",
                    "Sales_Rep", "Quantity", "Unit_Price",
                    "Discount", "Revenue", "Order_Status"]
    df_display = df[display_cols].copy()
    df_display["Date"] = df_display["Date"].dt.strftime("%Y-%m-%d")

    write_table_header(ws, 1, display_cols)

    for i, (_, row) in enumerate(df_display.iterrows()):
        write_data_row(ws, 2 + i, list(row), shade=i % 2 == 0)

    # Freeze top row
    ws.freeze_panes = "A2"
    autofit_cols(ws)

# ── Main Orchestrator ──────────────────────────────────────────
def generate_report(input_csv, output_path=None):
    print("\n" + "═" * 55)
    print("   AUTOMATED EXCEL REPORT GENERATOR")
    print("═" * 55)

    report_date = datetime.now().strftime("%B %d, %Y  %H:%M")
    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"Sales_Report_{ts}.xlsx"

    # Step 1: Load & clean
    df = load_and_clean(input_csv)

    # Step 2: Analyse
    kpis, monthly, by_product, by_region, by_rep, status_counts = run_analysis(df)

    # Step 3: Build workbook
    print("📝  Building Excel workbook...")
    wb = Workbook()

    sheets = {
        "Executive Summary":      wb.active,
        "Monthly Trends":         wb.create_sheet("Monthly Trends"),
        "Product Analysis":       wb.create_sheet("Product Analysis"),
        "Regional Analysis":      wb.create_sheet("Regional Analysis"),
        "Sales Rep Leaderboard":  wb.create_sheet("Sales Rep Leaderboard"),
        "Raw Data":               wb.create_sheet("Raw Data"),
    }
    wb.active.title = "Executive Summary"

    build_summary_sheet(sheets["Executive Summary"],    kpis, status_counts, report_date)
    build_monthly_sheet(sheets["Monthly Trends"],       monthly)
    build_product_sheet(sheets["Product Analysis"],     by_product)
    build_regional_sheet(sheets["Regional Analysis"],   by_region)
    build_leaderboard_sheet(sheets["Sales Rep Leaderboard"], by_rep)
    build_raw_data_sheet(sheets["Raw Data"],            df)

    # Step 4: Save
    wb.save(output_path)
    size_kb = round(os.path.getsize(output_path) / 1024, 1)
    print(f"\n✅  Report saved → {output_path}  ({size_kb} KB)")
    print("═" * 55 + "\n")
    return output_path

# ── Entry Point ────────────────────────────────────────────────
if __name__ == "__main__":
    csv_file = sys.argv[1] if len(sys.argv) > 1 else "sample_data.csv"
    generate_report(csv_file)
