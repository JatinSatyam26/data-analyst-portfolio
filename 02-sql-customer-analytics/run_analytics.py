"""
run_analytics.py
─────────────────────────────────────────────────────────────
SQL Customer Analytics Engine
Author : Jatin Prasad
Purpose: Runs a suite of advanced SQL queries against the
         customer analytics SQLite database and exports all
         results to a fully-formatted multi-sheet Excel report.

SQL Techniques Demonstrated:
  - CTEs (Common Table Expressions)
  - Window Functions (RANK, ROW_NUMBER, SUM OVER, AVG OVER)
  - Subqueries (correlated + non-correlated)
  - Aggregations with GROUP BY, HAVING
  - Multi-table JOINs
  - CASE WHEN logic
  - Date-based filtering and grouping
─────────────────────────────────────────────────────────────
"""

import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

DB_PATH = "customer_analytics.db"

# ── Colour Palette ────────────────────────────────────────────
DARK_BLUE   = "1B2A4A"
MID_BLUE    = "2E86C1"
LIGHT_BLUE  = "D6EAF8"
ACCENT      = "E67E22"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F4F6F7"
DARK_GRAY   = "2C3E50"

# ── Styling Helpers ───────────────────────────────────────────
def _side():
    return Side(style="thin", color="BBBBBB")

def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _hfont(size=10, color=WHITE):
    return Font(name="Calibri", size=size, bold=True, color=color)

def _bfont(size=10, bold=False, color=DARK_GRAY):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center")

def autofit(ws, min_w=10, max_w=45):
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value), default=8
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, min_w), max_w)

def write_sheet(ws, title, subtitle, df, show_index=False):
    """Write a complete styled sheet from a DataFrame."""
    ws.sheet_view.showGridLines = False

    # Title banner
    total_cols = len(df.columns) + (1 if show_index else 0)
    end_col    = get_column_letter(max(total_cols, 2))
    ws.merge_cells(f"A1:{end_col}1")
    c = ws["A1"]
    c.value     = title
    c.font      = Font(name="Calibri", size=13, bold=True, color=WHITE)
    c.fill      = _fill(DARK_BLUE)
    c.alignment = _center()
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f"A2:{end_col}2")
    s = ws["A2"]
    s.value     = subtitle
    s.font      = Font(name="Calibri", size=9, italic=True, color="888888")
    s.alignment = _center()
    ws.row_dimensions[2].height = 16

    # Column headers
    col_names = (["#"] if show_index else []) + list(df.columns)
    for ci, name in enumerate(col_names, start=1):
        cell = ws.cell(row=3, column=ci, value=name)
        cell.font      = _hfont()
        cell.fill      = _fill(MID_BLUE)
        cell.alignment = _center()
        cell.border    = _border()
    ws.row_dimensions[3].height = 20

    # Data rows
    for ri, (idx, row) in enumerate(df.iterrows()):
        bg     = LIGHT_BLUE if ri % 2 == 0 else WHITE
        values = ([ri + 1] if show_index else []) + list(row)
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=4 + ri, column=ci, value=val)
            cell.font      = _bfont()
            cell.fill      = _fill(bg)
            cell.alignment = _center()
            cell.border    = _border()

    ws.freeze_panes = "A4"
    autofit(ws)

# ── SQL Queries ───────────────────────────────────────────────
QUERIES = {}

# Q1: Total revenue, orders, and AOV by customer segment
QUERIES["Revenue by Segment"] = (
    """
    -- Aggregation + CASE WHEN + JOIN
    SELECT
        c.segment                                       AS Customer_Segment,
        COUNT(DISTINCT o.order_id)                      AS Total_Orders,
        COUNT(DISTINCT c.customer_id)                   AS Total_Customers,
        ROUND(SUM(oi.quantity * oi.unit_price * (1 - oi.discount)), 2)
                                                        AS Total_Revenue,
        ROUND(SUM(oi.quantity * oi.unit_price * (1 - oi.discount))
              / COUNT(DISTINCT o.order_id), 2)          AS Avg_Order_Value,
        ROUND(SUM(oi.quantity * oi.unit_price * (1 - oi.discount))
              / COUNT(DISTINCT c.customer_id), 2)       AS Revenue_Per_Customer
    FROM customers c
    JOIN orders o      ON c.customer_id = o.customer_id
    JOIN order_items oi ON o.order_id   = oi.order_id
    WHERE o.status = 'Completed'
    GROUP BY c.segment
    ORDER BY Total_Revenue DESC
    """,
    "Revenue by Segment",
    "Aggregation + multi-table JOIN | Shows revenue, orders, and AOV broken down by customer segment"
)

# Q2: Top 10 customers by lifetime value using CTE
QUERIES["Top Customers by LTV"] = (
    """
    -- CTE + Subquery + Window Function
    WITH customer_revenue AS (
        SELECT
            c.customer_id,
            c.name,
            c.segment,
            c.region,
            COUNT(DISTINCT o.order_id)                              AS Total_Orders,
            ROUND(SUM(oi.quantity * oi.unit_price * (1 - oi.discount)), 2) AS Lifetime_Value
        FROM customers c
        JOIN orders o       ON c.customer_id = o.customer_id
        JOIN order_items oi ON o.order_id    = oi.order_id
        WHERE o.status = 'Completed'
        GROUP BY c.customer_id
    ),
    ranked AS (
        SELECT *,
               RANK() OVER (ORDER BY Lifetime_Value DESC) AS LTV_Rank
        FROM customer_revenue
    )
    SELECT
        LTV_Rank      AS Rank,
        name          AS Customer_Name,
        segment       AS Segment,
        region        AS Region,
        Total_Orders,
        Lifetime_Value
    FROM ranked
    WHERE LTV_Rank <= 10
    ORDER BY LTV_Rank
    """,
    "Top 10 Customers — Lifetime Value",
    "CTE + RANK() window function | Identifies highest-value customers by total completed order revenue"
)

# Q3: Monthly revenue trend with running total (window function)
QUERIES["Monthly Revenue Trend"] = (
    """
    -- Window Function: SUM OVER (running total) + date grouping
    WITH monthly AS (
        SELECT
            STRFTIME('%Y-%m', o.order_date)                             AS Month,
            COUNT(DISTINCT o.order_id)                                  AS Orders,
            ROUND(SUM(oi.quantity * oi.unit_price * (1 - oi.discount)), 2) AS Monthly_Revenue
        FROM orders o
        JOIN order_items oi ON o.order_id = oi.order_id
        WHERE o.status = 'Completed'
        GROUP BY Month
    )
    SELECT
        Month,
        Orders,
        Monthly_Revenue,
        ROUND(SUM(Monthly_Revenue) OVER (ORDER BY Month), 2) AS Running_Total,
        ROUND(AVG(Monthly_Revenue) OVER (
            ORDER BY Month ROWS BETWEEN 2 PRECEDING AND CURRENT ROW
        ), 2)                                                  AS Rolling_3M_Avg
    FROM monthly
    ORDER BY Month
    """,
    "Monthly Revenue Trend",
    "SUM OVER + AVG OVER window functions | Monthly revenue with running total and 3-month rolling average"
)

# Q4: Product performance with profit margin
QUERIES["Product Performance"] = (
    """
    -- JOIN + calculated profit margin + ORDER BY
    SELECT
        p.product_name                                              AS Product,
        p.category                                                  AS Category,
        SUM(oi.quantity)                                            AS Units_Sold,
        ROUND(SUM(oi.quantity * oi.unit_price * (1 - oi.discount)), 2) AS Revenue,
        ROUND(SUM(oi.quantity * p.cost_price), 2)                  AS Total_Cost,
        ROUND(SUM(oi.quantity * oi.unit_price * (1 - oi.discount))
              - SUM(oi.quantity * p.cost_price), 2)                AS Gross_Profit,
        ROUND(((SUM(oi.quantity * oi.unit_price * (1 - oi.discount))
              - SUM(oi.quantity * p.cost_price))
              / SUM(oi.quantity * oi.unit_price * (1 - oi.discount))) * 100, 1)
                                                                   AS Profit_Margin_Pct
    FROM products p
    JOIN order_items oi ON p.product_id = oi.product_id
    JOIN orders o       ON oi.order_id  = o.order_id
    WHERE o.status = 'Completed'
    GROUP BY p.product_id
    ORDER BY Revenue DESC
    """,
    "Product Performance",
    "Multi-table JOIN + profit margin calculation | Revenue, cost, and gross profit per product"
)

# Q5: Customers who have NEVER placed a completed order (correlated subquery)
QUERIES["Inactive Customers"] = (
    """
    -- Correlated Subquery + NOT EXISTS
    SELECT
        c.customer_id   AS Customer_ID,
        c.name          AS Customer_Name,
        c.segment       AS Segment,
        c.region        AS Region,
        c.signup_date   AS Signup_Date,
        COALESCE(
            (SELECT COUNT(*) FROM orders o WHERE o.customer_id = c.customer_id),
            0
        )               AS Total_Orders_Placed,
        CASE
            WHEN NOT EXISTS (
                SELECT 1 FROM orders o
                WHERE o.customer_id = c.customer_id
                  AND o.status = 'Completed'
            ) THEN 'No Completed Orders'
            ELSE 'Has Completed Orders'
        END             AS Status
    FROM customers c
    WHERE NOT EXISTS (
        SELECT 1 FROM orders o
        WHERE o.customer_id = c.customer_id
          AND o.status = 'Completed'
    )
    ORDER BY c.signup_date
    """,
    "Inactive Customers",
    "Correlated subquery + NOT EXISTS | Identifies customers who signed up but never completed a purchase"
)

# Q6: Regional performance with rank (window function)
QUERIES["Regional Ranking"] = (
    """
    -- Window Function: RANK() OVER PARTITION
    WITH region_stats AS (
        SELECT
            c.region,
            c.segment,
            COUNT(DISTINCT o.order_id)                                  AS Orders,
            ROUND(SUM(oi.quantity * oi.unit_price * (1 - oi.discount)), 2) AS Revenue
        FROM customers c
        JOIN orders o       ON c.customer_id = o.customer_id
        JOIN order_items oi ON o.order_id    = oi.order_id
        WHERE o.status = 'Completed'
        GROUP BY c.region, c.segment
    )
    SELECT
        region          AS Region,
        segment         AS Segment,
        Orders,
        Revenue,
        RANK() OVER (PARTITION BY region ORDER BY Revenue DESC) AS Rank_Within_Region,
        RANK() OVER (ORDER BY Revenue DESC)                     AS Overall_Rank
    FROM region_stats
    ORDER BY region, Rank_Within_Region
    """,
    "Regional Ranking",
    "RANK() OVER PARTITION BY | Ranks customer segments within each region and globally"
)

# Q7: Repeat vs one-time buyers
QUERIES["Repeat vs One-Time Buyers"] = (
    """
    -- Subquery + CASE WHEN + aggregation
    WITH order_counts AS (
        SELECT
            customer_id,
            COUNT(order_id) AS completed_orders
        FROM orders
        WHERE status = 'Completed'
        GROUP BY customer_id
    )
    SELECT
        CASE
            WHEN oc.completed_orders = 1 THEN 'One-Time Buyer'
            WHEN oc.completed_orders BETWEEN 2 AND 4 THEN 'Repeat Buyer'
            ELSE 'Loyal Customer (5+)'
        END                     AS Buyer_Type,
        COUNT(*)                AS Customer_Count,
        ROUND(AVG(oc.completed_orders), 1) AS Avg_Orders,
        ROUND(COUNT(*) * 100.0 / SUM(COUNT(*)) OVER (), 1) AS Pct_of_Total
    FROM customers c
    JOIN order_counts oc ON c.customer_id = oc.customer_id
    GROUP BY Buyer_Type
    ORDER BY Avg_Orders DESC
    """,
    "Repeat vs One-Time Buyers",
    "CASE WHEN + window function for percentage | Segments customers by purchase frequency"
)

# ── Excel Builder ─────────────────────────────────────────────
def build_excel(results, output_path):
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    for sheet_name, (df, subtitle) in results.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        write_sheet(ws, sheet_name, subtitle, df, show_index=False)

    wb.save(output_path)
    size_kb = round(os.path.getsize(output_path) / 1024, 1)
    print(f"✅  Report saved → {output_path}  ({size_kb} KB)")

# ── Main ──────────────────────────────────────────────────────
def run():
    print("\n" + "═" * 58)
    print("   SQL CUSTOMER ANALYTICS ENGINE")
    print("═" * 58)

    conn = sqlite3.connect(DB_PATH)
    results = {}

    for name, (sql, sheet_title, subtitle) in QUERIES.items():
        print(f"⚙️   Running: {name}")
        df = pd.read_sql_query(sql, conn)
        results[sheet_title] = (df, subtitle)
        print(f"    → {len(df)} rows returned")

    conn.close()

    ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"Customer_Analytics_Report_{ts}.xlsx"
    print(f"\n📝  Building Excel report...")
    build_excel(results, output_path)
    print("═" * 58 + "\n")

if __name__ == "__main__":
    run()
