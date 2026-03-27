"""
data_quality_bot.py
─────────────────────────────────────────────────────────────
Data Quality & Validation Bot
Author : Jatin Prasad
Purpose: Automatically scans any CSV file for data quality
         issues and produces a detailed, styled Excel report
         with flagged rows, issue counts, and a health score.

Checks Performed:
  1. Missing Values         — nulls in every column
  2. Duplicate Rows         — fully duplicated records
  3. Negative Numerics      — negative values in numeric cols
  4. Invalid Emails         — regex-based email validation
  5. Out-of-Range Ages      — age outside [0, 120]
  6. Inconsistent Casing    — text columns with mixed case/whitespace
  7. Invalid Dates          — unparseable date strings
  8. Numeric Outliers       — IQR-based outlier detection

Output Sheets:
  - Health Dashboard   — overall data quality score + issue summary
  - Missing Values     — rows and columns with nulls
  - Duplicates         — duplicated rows
  - Negative Values    — rows with negative numerics
  - Invalid Emails     — rows with malformed emails
  - Age Violations     — rows with invalid ages
  - Casing Issues      — rows with inconsistent text
  - Invalid Dates      — rows with bad date formats
  - Outliers           — rows with statistical outliers
  - Cleaned Data       — data after auto-fixing common issues
─────────────────────────────────────────────────────────────
"""

import pandas as pd
import numpy as np
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import datetime
import sys
import os

# ── Palette ───────────────────────────────────────────────────
DARK_RED   = "922B21"
MID_RED    = "E74C3C"
LIGHT_RED  = "FADBD8"
DARK_GREEN = "1E8449"
MID_GREEN  = "27AE60"
LIGHT_GREEN= "D5F5E3"
GOLD       = "F39C12"
DARK_GRAY  = "2C3E50"
LIGHT_GRAY = "F4F6F7"
WHITE      = "FFFFFF"
DARK_BLUE  = "1B2A4A"
MID_BLUE   = "2E86C1"

EMAIL_REGEX = re.compile(r"^[\w\.\+\-]+@[\w\-]+\.[a-zA-Z]{2,}$")
DATE_COLS   = ["order_date","date","signup_date","created_at","updated_at"]
EMAIL_COLS  = ["email","email_address","contact_email"]
AGE_COLS    = ["age","customer_age","user_age"]

# ── Style Helpers ─────────────────────────────────────────────
def _border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(size=10, bold=False, color=DARK_GRAY, italic=False):
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def autofit(ws, min_w=10, max_w=50):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(max_len + 2, min_w), max_w)

def write_banner(ws, title, subtitle, span=8, bg=DARK_BLUE):
    ws.sheet_view.showGridLines = False
    end = get_column_letter(max(span, 2))
    ws.merge_cells(f"A1:{end}1")
    c = ws["A1"]
    c.value, c.font = title, _font(14, bold=True, color=WHITE)
    c.fill, c.alignment = _fill(bg), _center()
    ws.row_dimensions[1].height = 32
    ws.merge_cells(f"A2:{end}2")
    s = ws["A2"]
    s.value, s.font = subtitle, _font(9, italic=True, color="888888")
    s.alignment = _center()
    ws.row_dimensions[2].height = 16

def write_header_row(ws, row, cols, bg=MID_BLUE):
    for ci, name in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=name)
        c.font, c.fill = _font(10, bold=True, color=WHITE), _fill(bg)
        c.alignment, c.border = _center(), _border()
    ws.row_dimensions[row].height = 20

def write_data_row(ws, row_idx, values, shade=False):
    bg = LIGHT_RED if shade else WHITE
    for ci, val in enumerate(values, 1):
        c = ws.cell(row=row_idx, column=ci, value=val)
        c.font, c.fill = _font(), _fill(bg)
        c.alignment, c.border = _center(), _border()

# ── Checks ────────────────────────────────────────────────────
def check_missing(df):
    rows = []
    for col in df.columns:
        for idx in df[df[col].isnull()].index:
            rows.append({"Row": idx+2, "Column": col, "Issue": "Missing / Null value"})
    return pd.DataFrame(rows)

def check_duplicates(df):
    dupes = df[df.duplicated(keep=False)].copy()
    dupes.insert(0, "Row", dupes.index + 2)
    dupes.insert(1, "Issue", "Duplicate row")
    return dupes.reset_index(drop=True)

def check_negative_numerics(df):
    rows = []
    for col in df.select_dtypes(include=[np.number]).columns:
        for idx in df[df[col] < 0].index:
            rows.append({"Row": idx+2, "Column": col,
                         "Value": df.at[idx, col], "Issue": "Negative numeric value"})
    return pd.DataFrame(rows)

def check_emails(df):
    rows = []
    for col in EMAIL_COLS:
        if col not in df.columns:
            continue
        for idx, val in df[col].items():
            if pd.isnull(val):
                continue
            if not EMAIL_REGEX.match(str(val).strip()):
                rows.append({"Row": idx+2, "Column": col,
                             "Value": val, "Issue": "Invalid email format"})
    return pd.DataFrame(rows)

def check_age(df):
    rows = []
    for col in AGE_COLS:
        if col not in df.columns:
            continue
        for idx, val in df[col].items():
            if pd.isnull(val):
                continue
            try:
                if not (0 <= float(val) <= 120):
                    rows.append({"Row": idx+2, "Column": col,
                                 "Value": val, "Issue": "Age out of range [0–120]"})
            except (ValueError, TypeError):
                rows.append({"Row": idx+2, "Column": col,
                             "Value": val, "Issue": "Non-numeric age"})
    return pd.DataFrame(rows)

def check_casing(df):
    rows = []
    skip = set(EMAIL_COLS + DATE_COLS + ["status","order_status"])
    for col in df.select_dtypes(include=["object"]).columns:
        if col in skip:
            continue
        for idx, val in df[col].items():
            if pd.isnull(val):
                continue
            s = str(val)
            if s != s.strip() or (s != s.title() and s != s.upper() and s != s.lower()):
                rows.append({"Row": idx+2, "Column": col,
                             "Raw Value": repr(val), "Issue": "Inconsistent casing or whitespace"})
    return pd.DataFrame(rows)

def check_dates(df):
    rows = []
    for col in DATE_COLS:
        if col not in df.columns:
            continue
        for idx, val in df[col].items():
            if pd.isnull(val) or str(val).strip() == "":
                continue
            try:
                pd.to_datetime(str(val))
            except Exception:
                rows.append({"Row": idx+2, "Column": col,
                             "Value": val, "Issue": "Unparseable date format"})
    return pd.DataFrame(rows)

def check_outliers(df):
    rows = []
    skip = set(AGE_COLS + ["customer_id","id","order_id","quantity","qty"])
    for col in df.select_dtypes(include=[np.number]).columns:
        if col in skip:
            continue
        q1, q3 = df[col].quantile(0.25), df[col].quantile(0.75)
        iqr = q3 - q1
        lo, hi = q1 - 3*iqr, q3 + 3*iqr
        for idx in df[(df[col] < lo) | (df[col] > hi)].index:
            rows.append({"Row": idx+2, "Column": col,
                         "Value": df.at[idx, col],
                         "Lower Fence": round(lo, 2),
                         "Upper Fence": round(hi, 2),
                         "Issue": "Statistical outlier (3×IQR)"})
    return pd.DataFrame(rows)

def auto_clean(df):
    c = df.copy()
    c.drop_duplicates(inplace=True)
    for col in c.select_dtypes(include=["object"]).columns:
        c[col] = c[col].str.strip()
    for col in ["region","status","product","name"]:
        if col in c.columns:
            c[col] = c[col].str.title()
    for col in c.select_dtypes(include=np.number).columns:
        c = c[~(c[col] < 0)]
    return c.reset_index(drop=True)

# ── Dashboard ─────────────────────────────────────────────────
def build_dashboard(ws, df, summary, report_date):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value, c.font = "🤖  DATA QUALITY BOT — HEALTH DASHBOARD", _font(15, bold=True, color=WHITE)
    c.fill, c.alignment = _fill(DARK_BLUE), _center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value, s.font = f"Scan completed: {report_date}", _font(9, italic=True, color="888888")
    s.alignment = _center()

    total_cells  = len(df) * len(df.columns)
    total_issues = sum(summary.values())
    health_pct   = round((total_cells - total_issues) / total_cells * 100, 1) if total_cells else 0
    hcolor       = DARK_GREEN if health_pct >= 90 else (GOLD if health_pct >= 75 else DARK_RED)

    cards = [("Total Rows", len(df)), ("Total Columns", len(df.columns)),
             ("Total Issues", total_issues), ("Health Score", f"{health_pct}%")]
    for i, (label, val) in enumerate(cards):
        col = i*2 + 1
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
        lc = ws.cell(row=4, column=col, value=label)
        lc.font, lc.fill, lc.alignment = _font(9, bold=True, color=WHITE), _fill(MID_BLUE), _center()
        ws.row_dimensions[4].height = 18
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)
        vc = ws.cell(row=5, column=col, value=val)
        vc.font = _font(16, bold=True, color=hcolor if label == "Health Score" else DARK_GRAY)
        vc.fill, vc.alignment, vc.border = _fill(LIGHT_GRAY), _center(), _border()
        ws.row_dimensions[5].height = 32

    ws.merge_cells("A7:H7")
    h = ws["A7"]
    h.value, h.font = "ISSUE SUMMARY BY CATEGORY", _font(11, bold=True, color=WHITE)
    h.fill, h.alignment = _fill(DARK_RED), _center()

    write_header_row(ws, 8, ["Check", "Issues Found", "Severity", "Recommendation"], bg=MID_RED)
    sev = {
        "Missing Values":  ("🔴 High",   "Impute or drop affected rows"),
        "Duplicate Rows":  ("🔴 High",   "Remove before analysis"),
        "Negative Values": ("🔴 High",   "Validate data entry at source"),
        "Invalid Emails":  ("🟡 Medium", "Flag for re-collection"),
        "Age Violations":  ("🟡 Medium", "Investigate extreme values"),
        "Casing Issues":   ("🟢 Low",    "Standardise with str.strip() / str.title()"),
        "Invalid Dates":   ("🔴 High",   "Enforce ISO 8601: YYYY-MM-DD"),
        "Outliers":        ("🟡 Medium", "Investigate — may be valid or errors"),
    }
    for i, (check, count) in enumerate(summary.items()):
        severity, rec = sev.get(check, ("🟡 Medium", "Review manually"))
        write_data_row(ws, 9+i, [check, count, severity, rec], shade=i%2==0)

    # Chart
    chart = BarChart()
    chart.type, chart.title = "bar", "Issues by Category"
    chart.height, chart.width, chart.style = 12, 18, 10
    data_ref   = Reference(ws, min_col=2, min_row=8, max_row=8+len(summary))
    labels_ref = Reference(ws, min_col=1, min_row=9, max_row=8+len(summary))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(labels_ref)
    ws.add_chart(chart, "E7")
    for i in range(1, 9):
        ws.column_dimensions[get_column_letter(i)].width = 20

# ── Main ──────────────────────────────────────────────────────
def run_bot(csv_path, output_path=None):
    print("\n" + "═"*58)
    print("   🤖  DATA QUALITY & VALIDATION BOT")
    print("═"*58)

    report_date = datetime.now().strftime("%B %d, %Y  %H:%M")
    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"DQ_Report_{ts}.xlsx"

    print(f"📂  Loading: {csv_path}")
    df = pd.read_csv(csv_path)
    print(f"    {len(df)} rows × {len(df.columns)} columns")

    checks = {
        "Missing Values":  check_missing,
        "Duplicate Rows":  check_duplicates,
        "Negative Values": check_negative_numerics,
        "Invalid Emails":  check_emails,
        "Age Violations":  check_age,
        "Casing Issues":   check_casing,
        "Invalid Dates":   check_dates,
        "Outliers":        check_outliers,
    }

    results, summary = {}, {}
    for name, fn in checks.items():
        print(f"⚙️   Checking: {name}...")
        r = fn(df)
        results[name], summary[name] = r, len(r)
        print(f"    → {len(r)} issues found")

    cleaned = auto_clean(df)
    print(f"\n📊  Total issues: {sum(summary.values())}")
    print("📝  Building report...")

    wb = Workbook()
    wb.active.title = "Health Dashboard"
    build_dashboard(wb.active, df, summary, report_date)

    subtitles = {
        "Missing Values":  "Null or empty values detected across all columns",
        "Duplicate Rows":  "Fully duplicated records that should be removed before analysis",
        "Negative Values": "Negative numbers found in columns that should only be positive",
        "Invalid Emails":  "Email addresses that fail regex pattern validation",
        "Age Violations":  "Age values outside the valid human range [0–120]",
        "Casing Issues":   "Text with inconsistent capitalisation or extra whitespace",
        "Invalid Dates":   "Date strings that cannot be parsed into a valid date",
        "Outliers":        "Values more than 3×IQR beyond Q1/Q3 — likely data errors",
    }

    col_map = {
        "Missing Values":  ["Row","Column","Issue"],
        "Negative Values": ["Row","Column","Value","Issue"],
        "Invalid Emails":  ["Row","Column","Value","Issue"],
        "Age Violations":  ["Row","Column","Value","Issue"],
        "Casing Issues":   ["Row","Column","Raw Value","Issue"],
        "Invalid Dates":   ["Row","Column","Value","Issue"],
        "Outliers":        ["Row","Column","Value","Lower Fence","Upper Fence","Issue"],
    }

    for name, df_issues in results.items():
        ws = wb.create_sheet(title=name[:31])
        write_banner(ws, f"🔍  {name}", subtitles[name], span=8, bg=DARK_RED)

        if name == "Duplicate Rows":
            cols = ["Row","Issue"] + list(df.columns)
        else:
            cols = col_map[name]

        write_header_row(ws, 3, cols, bg=MID_RED)

        if df_issues.empty:
            ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=len(cols))
            c = ws.cell(row=4, column=1, value="✅  No issues found")
            c.font, c.alignment = _font(10, color=DARK_GREEN, bold=True), _center()
        else:
            for i, (_, row) in enumerate(df_issues.iterrows()):
                vals = [row.get(c, "") for c in cols]
                write_data_row(ws, 4+i, vals, shade=i%2==0)
        autofit(ws)

    # Cleaned sheet
    ws_c = wb.create_sheet("Cleaned Data")
    write_banner(ws_c, "✅  Auto-Cleaned Dataset",
                 f"Duplicates removed · Whitespace stripped · Negative values removed · {len(cleaned)} rows retained",
                 span=len(cleaned.columns), bg=DARK_GREEN)
    write_header_row(ws_c, 3, list(cleaned.columns), bg=MID_GREEN)
    for i, (_, row) in enumerate(cleaned.iterrows()):
        bg = LIGHT_GREEN if i%2==0 else WHITE
        for ci, val in enumerate(row, 1):
            c = ws_c.cell(row=4+i, column=ci, value=val)
            c.font, c.fill = _font(), _fill(bg)
            c.alignment, c.border = _center(), _border()
    ws_c.freeze_panes = "A4"
    autofit(ws_c)

    wb.save(output_path)
    size_kb = round(os.path.getsize(output_path)/1024, 1)
    print(f"\n✅  Report saved → {output_path}  ({size_kb} KB)")
    print(f"    Cleaned dataset: {len(cleaned)} rows retained from {len(df)} original")
    print("═"*58 + "\n")

if __name__ == "__main__":
    csv_file = sys.argv[1] if len(sys.argv) > 1 else "dirty_data.csv"
    run_bot(csv_file)
