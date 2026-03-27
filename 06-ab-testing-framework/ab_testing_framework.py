"""
ab_testing_framework.py
────────────────────────────────────────────────────────────────
A/B Testing Analysis Framework
Author : Jatin Prasad
Purpose: Runs a full statistical A/B test analysis on any
         binary metric dataset and exports results to a
         formatted multi-sheet Excel report.

Statistical Methods Used:
  - Two-proportion Z-test        (hypothesis testing)
  - Chi-squared test             (independence test)
  - 95% Confidence Intervals     (Wilson score method)
  - Relative & Absolute Uplift   (effect size)
  - Statistical Power            (sensitivity check)
  - Sample Size Calculator       (pre-test planning tool)
  - Daily conversion trend       (time-series breakdown)

Verdict Logic:
  - p < 0.05  AND  uplift > 0  → "✅ Variant Wins — Launch It"
  - p < 0.05  AND  uplift < 0  → "❌ Control Wins — Keep Original"
  - p >= 0.05                  → "⚠️  Inconclusive — Need More Data"
────────────────────────────────────────────────────────────────
"""

import pandas as pd
import numpy as np
from scipy.stats import chi2_contingency, norm
import math
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Palette ────────────────────────────────────────────────────
DARK_PURPLE  = "2C1654"
MID_PURPLE   = "7D3C98"
LIGHT_PURPLE = "E8DAEF"
GREEN_WIN    = "1E8449"
RED_LOSE     = "922B21"
ORANGE_INC   = "E67E22"
WHITE        = "FFFFFF"
DARK_GRAY    = "2C3E50"
ACCENT       = "F39C12"

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(size=10, bold=False, color=DARK_GRAY):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def autofit(ws, min_w=14, max_w=55):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, min_w), max_w)

def banner(ws, row, text, span, bg=DARK_PURPLE, size=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = _font(size=size, bold=True, color=WHITE)
    c.fill = _fill(bg)
    c.alignment = _center()
    ws.row_dimensions[row].height = 32

def subheader(ws, row, text, span):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = _font(size=10, bold=True, color=WHITE)
    c.fill = _fill(MID_PURPLE)
    c.alignment = _center()
    ws.row_dimensions[row].height = 20

def table_header(ws, row, cols):
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=col)
        c.font = _font(bold=True, color=WHITE)
        c.fill = _fill(MID_PURPLE)
        c.alignment = _center()
        c.border = _border()
    ws.row_dimensions[row].height = 18

def data_row(ws, row, values, shade=False, color=None):
    bg = color if color else (LIGHT_PURPLE if shade else WHITE)
    for ci, val in enumerate(values, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.font = _font()
        c.fill = _fill(bg)
        c.alignment = _center()
        c.border = _border()

# ── Wilson Confidence Interval ─────────────────────────────────
def wilson_ci(successes, n, z=1.96):
    if n == 0:
        return 0, 0
    p_hat  = successes / n
    denom  = 1 + z**2 / n
    centre = (p_hat + z**2 / (2*n)) / denom
    margin = (z * math.sqrt(p_hat*(1-p_hat)/n + z**2/(4*n**2))) / denom
    return max(0, centre - margin), min(1, centre + margin)

# ── Sample Size Calculator ─────────────────────────────────────
def required_sample_size(baseline_rate, mde, alpha=0.05, power=0.80):
    z_alpha = norm.ppf(1 - alpha / 2)
    z_beta  = norm.ppf(power)
    p1      = baseline_rate
    p2      = baseline_rate + mde
    p_bar   = (p1 + p2) / 2
    n = ((z_alpha * math.sqrt(2 * p_bar * (1 - p_bar)) +
          z_beta  * math.sqrt(p1*(1-p1) + p2*(1-p2))) ** 2) / (mde ** 2)
    return math.ceil(n)

# ── Core Analysis ──────────────────────────────────────────────
def analyse_test(df, metric_col, test_name):
    control = df[df["group"] == "Control"]
    variant = df[df["group"] == "Variant"]

    n_c   = len(control);  conv_c = control[metric_col].sum()
    n_v   = len(variant);  conv_v = variant[metric_col].sum()
    r_c   = conv_c / n_c;  r_v    = conv_v / n_v

    # Two-proportion Z-test
    p_pool  = (conv_c + conv_v) / (n_c + n_v)
    se      = math.sqrt(p_pool * (1 - p_pool) * (1/n_c + 1/n_v))
    z_stat  = (r_v - r_c) / se if se > 0 else 0
    p_value = 2 * (1 - norm.cdf(abs(z_stat)))

    # Chi-squared test
    contingency     = np.array([[conv_c, n_c - conv_c],
                                 [conv_v, n_v - conv_v]])
    chi2, p_chi2, _, _ = chi2_contingency(contingency)

    # Confidence intervals
    ci_c = wilson_ci(conv_c, n_c)
    ci_v = wilson_ci(conv_v, n_v)

    # Effect size
    abs_uplift = r_v - r_c
    rel_uplift = (r_v - r_c) / r_c * 100 if r_c > 0 else 0

    # Power
    power = 1 - norm.cdf(norm.ppf(0.975) - abs(z_stat))

    # Required sample size
    mde   = abs(abs_uplift) if abs(abs_uplift) > 0.001 else 0.01
    req_n = required_sample_size(r_c, mde)

    # Verdict
    sig = p_value < 0.05
    if sig and abs_uplift > 0:
        verdict       = "✅  Variant Wins — Launch It"
        verdict_color = GREEN_WIN
    elif sig and abs_uplift < 0:
        verdict       = "❌  Control Wins — Keep Original"
        verdict_color = RED_LOSE
    else:
        verdict       = "⚠️   Inconclusive — Collect More Data"
        verdict_color = ORANGE_INC

    results = {
        "test_name":       test_name,
        "metric_col":      metric_col,
        "control_name":    df[df["group"]=="Control"]["group_name"].iloc[0],
        "variant_name":    df[df["group"]=="Variant"]["group_name"].iloc[0],
        "n_control":       n_c,   "n_variant":    n_v,
        "conv_control":    int(conv_c), "conv_variant": int(conv_v),
        "rate_control":    r_c,   "rate_variant": r_v,
        "ci_c":            ci_c,  "ci_v":         ci_v,
        "abs_uplift":      abs_uplift,
        "rel_uplift":      rel_uplift,
        "z_stat":          z_stat,
        "p_value":         p_value,
        "chi2":            chi2,  "p_chi2":       p_chi2,
        "power":           power,
        "req_sample_size": req_n,
        "verdict":         verdict,
        "verdict_color":   verdict_color,
        "significant":     sig,
    }

    daily = (df.groupby(["day","group"])[metric_col]
               .agg(["sum","count"]).reset_index()
               .rename(columns={"sum":"conversions","count":"users"}))
    daily["rate"] = (daily["conversions"] / daily["users"]).round(4)

    return results, daily

# ── Test Sheet ─────────────────────────────────────────────────
def build_test_sheet(wb, results, daily, sheet_name):
    ws = wb.create_sheet(title=sheet_name[:31])
    ws.sheet_view.showGridLines = False
    r    = results
    SPAN = 6

    banner(ws, 1, f"A/B TEST — {r['test_name'].upper()}", SPAN)
    ws.merge_cells("A2:F2")
    s = ws["A2"]
    s.value     = f"Metric: {r['metric_col']}   |   Date: {datetime.now().strftime('%B %d, %Y')}"
    s.font      = _font(size=9, color="888888")
    s.alignment = _center()

    ws.merge_cells("A4:F4")
    v = ws["A4"]
    v.value     = r["verdict"]
    v.font      = Font(name="Calibri", size=15, bold=True, color=WHITE)
    v.fill      = _fill(r["verdict_color"])
    v.alignment = _center()
    ws.row_dimensions[4].height = 38

    subheader(ws, 6, "GROUP PERFORMANCE SUMMARY", SPAN)
    table_header(ws, 7, ["Group","Variant Name","Users","Conversions","Conv. Rate","95% CI"])
    data_row(ws, 8, ["Control", r["control_name"], f"{r['n_control']:,}",
                      f"{r['conv_control']:,}", f"{r['rate_control']:.2%}",
                      f"{r['ci_c'][0]:.2%} – {r['ci_c'][1]:.2%}"], shade=True)
    data_row(ws, 9, ["Variant", r["variant_name"], f"{r['n_variant']:,}",
                      f"{r['conv_variant']:,}", f"{r['rate_variant']:.2%}",
                      f"{r['ci_v'][0]:.2%} – {r['ci_v'][1]:.2%}"])

    subheader(ws, 11, "STATISTICAL TEST RESULTS", SPAN)
    table_header(ws, 12, ["Metric","Value","Interpretation","","",""])
    stats_rows = [
        ("Absolute Uplift",        f"{r['abs_uplift']:+.2%}",
         "Difference in conversion rate (Variant − Control)"),
        ("Relative Uplift",        f"{r['rel_uplift']:+.1f}%",
         "% improvement of Variant over Control"),
        ("Z-Statistic",            f"{r['z_stat']:.4f}",
         "Standard deviations from the null hypothesis"),
        ("P-Value (Z-test)",       f"{r['p_value']:.4f}",
         "< 0.05 = statistically significant result"),
        ("Chi-Squared Statistic",  f"{r['chi2']:.4f}",
         "Independence test — confirms Z-test"),
        ("P-Value (Chi-sq)",       f"{r['p_chi2']:.4f}",
         "Should align with Z-test p-value"),
        ("Statistical Power",      f"{r['power']:.1%}",
         "Probability of detecting a real effect (target ≥ 80%)"),
        ("Significant? (α=0.05)",  "YES ✅" if r["significant"] else "NO ❌",
         "Based on α = 0.05 threshold"),
        ("Required Sample / Group",f"{r['req_sample_size']:,}",
         "Min users per group needed to detect this effect size"),
    ]
    for i, (metric, value, interp) in enumerate(stats_rows):
        shade = i % 2 == 0
        bg    = LIGHT_PURPLE if shade else WHITE
        ws.cell(row=13+i, column=1, value=metric).fill   = _fill(bg)
        ws.cell(row=13+i, column=1).font                 = _font(bold=True)
        ws.cell(row=13+i, column=1).border               = _border()
        ws.cell(row=13+i, column=1).alignment            = _left()
        ws.cell(row=13+i, column=2, value=value).fill    = _fill(bg)
        ws.cell(row=13+i, column=2).font                 = Font(name="Calibri", size=10,
                                                                 bold=True, color=ACCENT)
        ws.cell(row=13+i, column=2).border               = _border()
        ws.cell(row=13+i, column=2).alignment            = _center()
        ws.merge_cells(start_row=13+i, start_column=3, end_row=13+i, end_column=6)
        ws.cell(row=13+i, column=3, value=interp).fill   = _fill(bg)
        ws.cell(row=13+i, column=3).font                 = _font(color="555555")
        ws.cell(row=13+i, column=3).border               = _border()
        ws.cell(row=13+i, column=3).alignment            = _left()

    start_r = 13 + len(stats_rows) + 2
    subheader(ws, start_r, "DAILY CONVERSION RATE TREND", SPAN)
    table_header(ws, start_r+1, ["Day","Control Users","Control Conv.",
                                   "Control Rate","Variant Users","Variant Conv.","Variant Rate"])
    ctrl_d = daily[daily["group"]=="Control"].set_index("day")
    var_d  = daily[daily["group"]=="Variant"].set_index("day")
    for i, day in enumerate(sorted(set(ctrl_d.index)|set(var_d.index))):
        c = ctrl_d.loc[day] if day in ctrl_d.index else None
        v = var_d.loc[day]  if day in var_d.index  else None
        data_row(ws, start_r+2+i, [
            f"Day {day}",
            f"{int(c['users']):,}"       if c is not None else "—",
            f"{int(c['conversions']):,}" if c is not None else "—",
            f"{c['rate']:.2%}"           if c is not None else "—",
            f"{int(v['users']):,}"       if v is not None else "—",
            f"{int(v['conversions']):,}" if v is not None else "—",
            f"{v['rate']:.2%}"           if v is not None else "—",
        ], shade=i%2==0)

    ws.freeze_panes = "A5"
    autofit(ws)

# ── Summary Sheet ──────────────────────────────────────────────
def build_summary_sheet(wb, all_results):
    ws = wb.create_sheet("📋 Executive Summary", 0)
    ws.sheet_view.showGridLines = False
    SPAN = 8

    banner(ws, 1, "A/B TESTING FRAMEWORK — EXECUTIVE SUMMARY", SPAN, size=14)
    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value     = f"Author: Jatin Prasad  |  Date: {datetime.now().strftime('%B %d, %Y')}  |  Tests: {len(all_results)}"
    s.font      = _font(size=9, color="888888")
    s.alignment = _center()

    table_header(ws, 4, ["Test Name","Metric","Control Rate","Variant Rate",
                           "Rel. Uplift","P-Value","Significant?","Verdict"])
    for i, r in enumerate(all_results):
        row_color = ("D5F5E3" if r["significant"] and r["abs_uplift"] > 0 else
                     "FADBD8" if r["significant"] and r["abs_uplift"] < 0 else "FEF9E7")
        data_row(ws, 5+i, [
            r["test_name"], r["metric_col"],
            f"{r['rate_control']:.2%}", f"{r['rate_variant']:.2%}",
            f"{r['rel_uplift']:+.1f}%", f"{r['p_value']:.4f}",
            "YES ✅" if r["significant"] else "NO ❌",
            r["verdict"].split("—")[0].strip()
        ], color=row_color)

    notes_start = 5 + len(all_results) + 2
    subheader(ws, notes_start, "METHODOLOGY", SPAN)
    notes = [
        ("Significance Level", "α = 0.05  (5% false positive rate)"),
        ("Hypothesis Test",    "Two-proportion Z-test (primary) + Chi-squared (confirmation)"),
        ("Confidence Intervals","Wilson score method at 95% confidence"),
        ("Power Target",       "80% — standard industry threshold"),
        ("Test Duration",      "14 days per test"),
    ]
    for i, (label, note) in enumerate(notes):
        bg = LIGHT_PURPLE if i%2==0 else WHITE
        ws.cell(row=notes_start+1+i, column=1, value=label).fill   = _fill(bg)
        ws.cell(row=notes_start+1+i, column=1).font                = _font(bold=True)
        ws.cell(row=notes_start+1+i, column=1).border              = _border()
        ws.cell(row=notes_start+1+i, column=1).alignment           = _left()
        ws.merge_cells(start_row=notes_start+1+i, start_column=2,
                       end_row=notes_start+1+i, end_column=8)
        ws.cell(row=notes_start+1+i, column=2, value=note).fill    = _fill(bg)
        ws.cell(row=notes_start+1+i, column=2).font                = _font()
        ws.cell(row=notes_start+1+i, column=2).border              = _border()
        ws.cell(row=notes_start+1+i, column=2).alignment           = _left()

    autofit(ws)

# ── Sample Size Calculator Sheet ───────────────────────────────
def build_calculator_sheet(wb):
    ws = wb.create_sheet("🧮 Sample Size Calculator")
    ws.sheet_view.showGridLines = False
    SPAN = 5

    banner(ws, 1, "SAMPLE SIZE CALCULATOR — PRE-TEST PLANNING", SPAN)
    ws.merge_cells("A2:E2")
    s = ws["A2"]
    s.value     = "How many users do you need per group before launching your A/B test?"
    s.font      = _font(size=9, color="888888")
    s.alignment = _center()

    subheader(ws, 4, "REQUIRED SAMPLE SIZE PER GROUP  (α = 0.05,  Power = 80%)", SPAN)
    table_header(ws, 5, ["Baseline Conv. Rate","MDE +1%","MDE +2%","MDE +3%","MDE +5%"])

    for i, base in enumerate([0.05, 0.10, 0.15, 0.20, 0.25, 0.30, 0.40, 0.50]):
        data_row(ws, 6+i, [
            f"{base:.0%}",
            *[f"{required_sample_size(base, mde):,}" for mde in [0.01, 0.02, 0.03, 0.05]]
        ], shade=i%2==0)

    subheader(ws, 15, "HOW TO USE THIS TABLE", SPAN)
    notes = [
        ("Baseline Rate",  "Your current conversion rate before running the test"),
        ("MDE",            "Minimum Detectable Effect — smallest lift you want to detect"),
        ("Sample Size",    "Users needed IN EACH GROUP (multiply by 2 for total traffic)"),
        ("Example",        "Baseline=10%, want to detect +2% lift → need 3,474 users per group (6,948 total)"),
    ]
    for i, (label, note) in enumerate(notes):
        bg = LIGHT_PURPLE if i%2==0 else WHITE
        ws.cell(row=16+i, column=1, value=label).fill   = _fill(bg)
        ws.cell(row=16+i, column=1).font                = _font(bold=True)
        ws.cell(row=16+i, column=1).border              = _border()
        ws.cell(row=16+i, column=1).alignment           = _left()
        ws.merge_cells(start_row=16+i, start_column=2, end_row=16+i, end_column=5)
        ws.cell(row=16+i, column=2, value=note).fill    = _fill(bg)
        ws.cell(row=16+i, column=2).font                = _font()
        ws.cell(row=16+i, column=2).border              = _border()
        ws.cell(row=16+i, column=2).alignment           = _left()

    autofit(ws)

# ── Main ───────────────────────────────────────────────────────
def run():
    print("\n" + "═"*58)
    print("   A/B TESTING ANALYSIS FRAMEWORK")
    print("═"*58)

    tests = [
        ("test1_email.csv",   "opened",    "Email Subject Line Test"),
        ("test2_button.csv",  "clicked",   "Landing Page Button Colour Test"),
        ("test3_pricing.csv", "converted", "Pricing Page Layout Test"),
    ]

    wb = Workbook()
    wb.remove(wb.active)
    all_results = []

    for csv_file, metric_col, test_name in tests:
        print(f"\n⚙️   Analysing: {test_name}")
        df = pd.read_csv(csv_file)
        results, daily = analyse_test(df, metric_col, test_name)
        all_results.append(results)
        short = test_name.replace(" Test","")[:28]
        build_test_sheet(wb, results, daily, short)
        print(f"    Control: {results['rate_control']:.2%}  |  Variant: {results['rate_variant']:.2%}")
        print(f"    P-value: {results['p_value']:.4f}  |  {results['verdict']}")

    build_summary_sheet(wb, all_results)
    build_calculator_sheet(wb)

    ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"AB_Test_Report_{ts}.xlsx"
    wb.save(output_path)
    size_kb = round(os.path.getsize(output_path)/1024, 1)
    print(f"\n✅  Report saved → {output_path}  ({size_kb} KB)")
    print("═"*58 + "\n")

if __name__ == "__main__":
    run()
