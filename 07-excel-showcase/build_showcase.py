"""
build_showcase.py
────────────────────────────────────────────────────────────────
Excel Advanced Functions Showcase
Author : Jatin Prasad
Purpose: Generates a fully working Excel workbook demonstrating
         advanced Excel skills used by professional Data Analysts.

Sheets:
  1. 🏠 Overview          — What each sheet demonstrates
  2. 📦 Product Lookup     — XLOOKUP, INDEX-MATCH, IFERROR
  3. 📊 Sales Summary      — SUMIFS, COUNTIFS, AVERAGEIFS
  4. 🔢 Nested Logic       — Nested IF, IFS, SWITCH, CHOOSE
  5. 📅 Date Intelligence  — EOMONTH, NETWORKDAYS, DATEDIF, WEEKDAY
  6. 📋 Dynamic Arrays     — UNIQUE, SORT, FILTER, SEQUENCE
  7. 🎨 Conditional Format — Rules, colour scales, icon sets
  8. ✅ Data Validation    — Dropdowns, numeric rules, custom rules
  9. 📖 Formula Reference  — Cheatsheet of every formula used

Every formula cell contains a REAL working Excel formula so the
reviewer can open the file and see live calculations.
────────────────────────────────────────────────────────────────
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import (
    ColorScaleRule, DataBarRule, IconSetRule,
    CellIsRule, FormulaRule
)
from openpyxl.styles.differential import DifferentialStyle
from datetime import datetime, date
import os

# ── Palette ────────────────────────────────────────────────────
NAVY        = "1B2A4A"
DARK_TEAL   = "0E6655"
MID_TEAL    = "1ABC9C"
LIGHT_TEAL  = "D1F2EB"
PURPLE      = "6C3483"
LIGHT_PURP  = "E8DAEF"
ORANGE      = "D35400"
LIGHT_ORANG = "FDEBD0"
BLUE        = "1A5276"
LIGHT_BLUE  = "D6EAF8"
GOLD        = "B7950B"
LIGHT_GOLD  = "FEF9E7"
RED         = "922B21"
LIGHT_RED   = "FADBD8"
GREEN       = "1E8449"
LIGHT_GREEN = "D5F5E3"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F4F6F7"
DARK_GRAY   = "2C3E50"
MID_GRAY    = "85929E"

# ── Style Helpers ──────────────────────────────────────────────
def _s(color="BBBBBB", style="thin"):
    return Side(style=style, color=color)

def _border(c="CCCCCC"):
    s = _s(c)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(h):
    return PatternFill("solid", fgColor=h)

def _font(size=10, bold=False, color=DARK_GRAY, italic=False):
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _right():
    return Alignment(horizontal="right", vertical="center")

def autofit(ws, min_w=10, max_w=55):
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, min_w), max_w)

def banner(ws, row, text, span, bg=NAVY, size=13, height=32):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = _font(size=size, bold=True, color=WHITE)
    c.fill      = _fill(bg)
    c.alignment = _center()
    ws.row_dimensions[row].height = height

def subhead(ws, row, text, span, bg=DARK_TEAL, size=11):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = _font(size=size, bold=True, color=WHITE)
    c.fill      = _fill(bg)
    c.alignment = _center()
    ws.row_dimensions[row].height = 22

def thead(ws, row, cols, bg=DARK_TEAL):
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=col)
        c.font      = _font(bold=True, color=WHITE, size=10)
        c.fill      = _fill(bg)
        c.alignment = _center()
        c.border    = _border()
    ws.row_dimensions[row].height = 18

def cell(ws, row, col, value, bg=WHITE, bold=False, italic=False,
         color=DARK_GRAY, size=10, align="center", border=True, height=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(size=size, bold=bold, color=color, italic=italic)
    c.fill      = _fill(bg)
    c.alignment = _center() if align == "center" else _left() if align == "left" else _right()
    if border:
        c.border = _border()
    if height:
        ws.row_dimensions[row].height = height
    return c

def formula_cell(ws, row, col, formula, bg=LIGHT_GOLD, bold=False):
    """Write a real Excel formula — visible as a calculated value when opened."""
    c = ws.cell(row=row, column=col, value=formula)
    c.font      = _font(bold=bold, color=DARK_GRAY)
    c.fill      = _fill(bg)
    c.alignment = _center()
    c.border    = _border()
    return c

def label_cell(ws, row, col, text, bg=LIGHT_GRAY):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = _font(italic=True, color=MID_GRAY, size=9)
    c.fill      = _fill(bg)
    c.alignment = _left()
    c.border    = _border()
    return c

# ══════════════════════════════════════════════════════════════
#  SHEET 1 — OVERVIEW
# ══════════════════════════════════════════════════════════════
def build_overview(ws):
    ws.sheet_view.showGridLines = False
    banner(ws, 1, "📊  Excel Advanced Functions Showcase", 4, bg=NAVY, size=14, height=38)

    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value     = "Author: Jatin Prasad  |  MS Robotics & Autonomous Systems  |  ASU Ira A. Fulton Schools of Engineering"
    c.font      = _font(size=9, italic=True, color=MID_GRAY)
    c.alignment = _center()

    ws.merge_cells("A3:D3")
    d = ws["A3"]
    d.value     = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
    d.font      = _font(size=9, italic=True, color=MID_GRAY)
    d.alignment = _center()

    subhead(ws, 5, "WORKBOOK CONTENTS", 4)
    thead(ws, 6, ["Sheet", "Topic", "Functions Demonstrated", "Skill Level"])

    sheets_info = [
        ("📦 Product Lookup",    "Lookup & Reference",  "XLOOKUP, INDEX-MATCH, IFERROR, VLOOKUP",               "⭐⭐⭐"),
        ("📊 Sales Summary",     "Conditional Aggregation", "SUMIFS, COUNTIFS, AVERAGEIFS, SUMPRODUCT",          "⭐⭐⭐"),
        ("🔢 Nested Logic",      "Logical Functions",   "Nested IF, IFS, SWITCH, CHOOSE, AND, OR",              "⭐⭐⭐"),
        ("📅 Date Intelligence", "Date & Time",         "EOMONTH, NETWORKDAYS, DATEDIF, WEEKDAY, TODAY",        "⭐⭐⭐"),
        ("📋 Dynamic Arrays",    "Modern Excel",        "UNIQUE, SORT, FILTER, SEQUENCE, XLOOKUP arrays",       "⭐⭐⭐⭐"),
        ("🎨 Cond. Formatting",  "Visual Analytics",    "Color scales, data bars, icon sets, formula rules",    "⭐⭐⭐"),
        ("✅ Data Validation",   "Data Integrity",      "Dropdown lists, numeric rules, custom formulas",       "⭐⭐⭐"),
        ("📖 Formula Reference", "Cheatsheet",          "All formulas used in this workbook with explanations", "Reference"),
    ]

    bg_cycle = [LIGHT_TEAL, WHITE]
    for i, (sheet, topic, funcs, level) in enumerate(sheets_info):
        bg = bg_cycle[i % 2]
        for ci, val in enumerate([sheet, topic, funcs, level], 1):
            c = ws.cell(row=7+i, column=ci, value=val)
            c.font      = _font(bold=(ci==1))
            c.fill      = _fill(bg)
            c.alignment = _center() if ci in [1,4] else _left()
            c.border    = _border()

    ws.merge_cells("A16:D16")
    note = ws["A16"]
    note.value = (
        "💡  HOW TO USE THIS WORKBOOK:  Every yellow-highlighted cell contains a real, working Excel formula. "
        "Click any yellow cell to see the formula in the formula bar. "
        "The 📖 Formula Reference sheet lists every formula used with plain-English explanations."
    )
    note.font      = _font(size=10, italic=True, color=BLUE)
    note.fill      = _fill(LIGHT_BLUE)
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    note.border    = _border()
    ws.row_dimensions[16].height = 50

    for c, w in [(1,22),(2,22),(3,50),(4,14)]:
        ws.column_dimensions[get_column_letter(c)].width = w


# ══════════════════════════════════════════════════════════════
#  SHEET 2 — PRODUCT LOOKUP  (XLOOKUP, INDEX-MATCH, IFERROR)
# ══════════════════════════════════════════════════════════════
def build_lookup(ws):
    ws.sheet_view.showGridLines = False
    banner(ws, 1, "📦  Lookup & Reference Functions — XLOOKUP · INDEX-MATCH · IFERROR", 7, bg=DARK_TEAL)

    # ── Product master table ──────────────────────────────────
    subhead(ws, 3, "PRODUCT MASTER TABLE  (Source Data)", 5, bg=PURPLE)
    thead(ws, 4, ["Product ID", "Product Name", "Category", "Unit Price", "Stock Qty"], bg=PURPLE)

    products = [
        ("P001","Laptop Pro 15",       "Electronics", 1299.99, 45),
        ("P002","Wireless Mouse",      "Accessories",   45.99, 320),
        ("P003","Mechanical Keyboard", "Accessories",   89.99, 185),
        ("P004","4K Monitor",          "Electronics",  399.99,  67),
        ("P005","USB-C Hub",           "Accessories",   59.99, 240),
        ("P006","Webcam HD",           "Electronics",   79.99, 110),
        ("P007","Ergonomic Chair",     "Furniture",    349.99,  22),
        ("P008","Standing Desk",       "Furniture",    599.99,  14),
        ("P009","Noise-Cancel Headset","Electronics",  249.99,  88),
        ("P010","Laptop Stand",        "Accessories",   39.99, 175),
    ]

    bg_cycle = [LIGHT_PURP, WHITE]
    for i, row in enumerate(products):
        bg = bg_cycle[i % 2]
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=5+i, column=ci, value=val)
            c.font      = _font()
            c.fill      = _fill(bg)
            c.alignment = _center()
            c.border    = _border()

    # ── XLOOKUP demo ─────────────────────────────────────────
    subhead(ws, 16, "XLOOKUP — Look up a product by ID", 7, bg=DARK_TEAL)
    thead(ws, 17, ["Lookup ID", "→ Product Name", "→ Category", "→ Price", "→ Stock", "Formula Used", ""], bg=DARK_TEAL)

    lookup_ids = ["P003", "P007", "P010", "P999"]
    for i, pid in enumerate(lookup_ids):
        r = 18 + i
        cell(ws, r, 1, pid, bg=LIGHT_TEAL, bold=True)
        # XLOOKUP formulas referencing the master table
        formula_cell(ws, r, 2, f'=IFERROR(XLOOKUP(A{r},A$5:A$14,B$5:B$14),"❌ Not Found")')
        formula_cell(ws, r, 3, f'=IFERROR(XLOOKUP(A{r},A$5:A$14,C$5:C$14),"❌ Not Found")')
        formula_cell(ws, r, 4, f'=IFERROR(XLOOKUP(A{r},A$5:A$14,D$5:D$14),"❌ Not Found")')
        formula_cell(ws, r, 5, f'=IFERROR(XLOOKUP(A{r},A$5:A$14,E$5:E$14),"❌ Not Found")')
        label_cell(ws, r, 6, f'=IFERROR(XLOOKUP(lookup, id_range, return_range), "Not Found")')

    # ── INDEX-MATCH demo ──────────────────────────────────────
    subhead(ws, 23, "INDEX-MATCH — Alternative to VLOOKUP (works left-to-right AND right-to-left)", 7, bg=BLUE)
    thead(ws, 24, ["Lookup Name", "→ Product ID", "→ Price", "→ Stock", "", "Formula Used", ""], bg=BLUE)

    names = ["Webcam HD", "Standing Desk", "Non Existent"]
    for i, name in enumerate(names):
        r = 25 + i
        cell(ws, r, 1, name, bg=LIGHT_BLUE, bold=True)
        formula_cell(ws, r, 2, f'=IFERROR(INDEX(A$5:A$14,MATCH(A{r},B$5:B$14,0)),"❌ Not Found")', bg=LIGHT_BLUE)
        formula_cell(ws, r, 3, f'=IFERROR(INDEX(D$5:D$14,MATCH(A{r},B$5:B$14,0)),"❌ Not Found")', bg=LIGHT_BLUE)
        formula_cell(ws, r, 4, f'=IFERROR(INDEX(E$5:E$14,MATCH(A{r},B$5:B$14,0)),"❌ Not Found")', bg=LIGHT_BLUE)
        label_cell(ws, r, 6, '=IFERROR(INDEX(return_col, MATCH(lookup, lookup_col, 0)), "Not Found")')

    for c, w in [(1,14),(2,24),(3,16),(4,12),(5,12),(6,52),(7,10)]:
        ws.column_dimensions[get_column_letter(c)].width = w


# ══════════════════════════════════════════════════════════════
#  SHEET 3 — SALES SUMMARY  (SUMIFS, COUNTIFS, AVERAGEIFS)
# ══════════════════════════════════════════════════════════════
def build_sales_summary(ws):
    ws.sheet_view.showGridLines = False
    banner(ws, 1, "📊  Conditional Aggregation — SUMIFS · COUNTIFS · AVERAGEIFS · SUMPRODUCT", 8, bg=DARK_TEAL)

    # Raw sales data
    subhead(ws, 3, "RAW SALES DATA  (Source)", 7, bg=ORANGE)
    thead(ws, 4, ["Order ID","Product","Region","Rep","Qty","Revenue","Status"], bg=ORANGE)

    sales = [
        ("O001","Laptop Pro 15",  "North","Alice",  2, 2599.98,"Completed"),
        ("O002","Webcam HD",      "South","Bob",    5,  399.95,"Completed"),
        ("O003","4K Monitor",     "North","Alice",  1,  399.99,"Returned"),
        ("O004","Ergonomic Chair","East", "Carol",  3, 1049.97,"Completed"),
        ("O005","Laptop Pro 15",  "West", "David",  1, 1299.99,"Completed"),
        ("O006","USB-C Hub",      "South","Bob",    8,  479.92,"Pending"),
        ("O007","Webcam HD",      "North","Alice",  2,  159.98,"Completed"),
        ("O008","Standing Desk",  "East", "Carol",  1,  599.99,"Completed"),
        ("O009","Laptop Pro 15",  "South","Bob",    2, 2599.98,"Completed"),
        ("O010","Ergonomic Chair","North","Alice",  2,  699.98,"Completed"),
        ("O011","4K Monitor",     "West", "David",  3, 1199.97,"Completed"),
        ("O012","USB-C Hub",      "East", "Carol",  5,  299.95,"Returned"),
        ("O013","Webcam HD",      "West", "David",  4,  319.96,"Completed"),
        ("O014","Laptop Pro 15",  "East", "Carol",  1, 1299.99,"Pending"),
        ("O015","Standing Desk",  "South","Bob",    2, 1199.98,"Completed"),
    ]

    bg_cycle = [LIGHT_ORANG, WHITE]
    for i, row in enumerate(sales):
        bg = bg_cycle[i % 2]
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=5+i, column=ci, value=val)
            c.font      = _font()
            c.fill      = _fill(bg)
            c.alignment = _center()
            c.border    = _border()

    data_end = 5 + len(sales) - 1  # row 19

    # SUMIFS
    subhead(ws, 22, "SUMIFS — Sum revenue by region AND status", 8, bg=DARK_TEAL)
    thead(ws, 23, ["Region", "Status", "Total Revenue", "Formula", "", "", "", ""], bg=DARK_TEAL)
    sumifs_cases = [
        ("North", "Completed"),
        ("South", "Completed"),
        ("East",  "Completed"),
        ("West",  "Completed"),
    ]
    for i, (region, status) in enumerate(sumifs_cases):
        r = 24 + i
        cell(ws, r, 1, region, bg=LIGHT_TEAL, bold=True)
        cell(ws, r, 2, status, bg=LIGHT_TEAL)
        formula_cell(ws, r, 3, f'=SUMIFS(F$5:F${data_end},C$5:C${data_end},A{r},G$5:G${data_end},B{r})')
        label_cell(ws, r, 4, f'=SUMIFS(revenue_range, region_range, "{region}", status_range, "{status}")')

    # COUNTIFS
    subhead(ws, 29, "COUNTIFS — Count orders by rep AND status", 8, bg=BLUE)
    thead(ws, 30, ["Sales Rep", "Status", "Order Count", "Formula", "", "", "", ""], bg=BLUE)
    countifs_cases = [
        ("Alice","Completed"), ("Bob","Completed"), ("Carol","Completed"), ("David","Completed")
    ]
    for i, (rep, status) in enumerate(countifs_cases):
        r = 31 + i
        cell(ws, r, 1, rep,    bg=LIGHT_BLUE, bold=True)
        cell(ws, r, 2, status, bg=LIGHT_BLUE)
        formula_cell(ws, r, 3, f'=COUNTIFS(D$5:D${data_end},A{r},G$5:G${data_end},B{r})', bg=LIGHT_BLUE)
        label_cell(ws, r, 4, f'=COUNTIFS(rep_range, "{rep}", status_range, "{status}")')

    # AVERAGEIFS + SUMPRODUCT
    subhead(ws, 36, "AVERAGEIFS & SUMPRODUCT — Average order value by region", 8, bg=GREEN)
    thead(ws, 37, ["Region", "Avg Revenue (Completed)", "Weighted Avg (SUMPRODUCT)", "Formula", "", "", "", ""], bg=GREEN)
    for i, region in enumerate(["North","South","East","West"]):
        r = 38 + i
        cell(ws, r, 1, region, bg=LIGHT_GREEN, bold=True)
        formula_cell(ws, r, 2, f'=IFERROR(AVERAGEIFS(F$5:F${data_end},C$5:C${data_end},A{r},G$5:G${data_end},"Completed"),0)', bg=LIGHT_GREEN)
        formula_cell(ws, r, 3, f'=IFERROR(SUMPRODUCT((C$5:C${data_end}=A{r})*(G$5:G${data_end}="Completed")*F$5:F${data_end})/SUMPRODUCT((C$5:C${data_end}=A{r})*(G$5:G${data_end}="Completed")*1),0)', bg=LIGHT_GREEN)
        label_cell(ws, r, 4, '=SUMPRODUCT((region=X)*(status="Completed")*revenue) / SUMPRODUCT((region=X)*(status="Completed"))')

    for c, w in [(1,16),(2,16),(3,22),(4,22),(5,22),(6,14),(7,14),(8,10)]:
        ws.column_dimensions[get_column_letter(c)].width = w


# ══════════════════════════════════════════════════════════════
#  SHEET 4 — NESTED LOGIC
# ══════════════════════════════════════════════════════════════
def build_nested_logic(ws):
    ws.sheet_view.showGridLines = False
    banner(ws, 1, "🔢  Logical Functions — Nested IF · IFS · SWITCH · AND · OR", 6, bg=PURPLE)

    # Source: scores and sales data
    subhead(ws, 3, "SAMPLE DATA", 6, bg=PURPLE)
    thead(ws, 4, ["Employee","Score","Revenue","Region","Dept",""], bg=PURPLE)
    people = [
        ("Alice",  92, 45200, "North","Sales"),
        ("Bob",    74, 28100, "South","Sales"),
        ("Carol",  85, 62400, "East", "Marketing"),
        ("David",  61, 15300, "West", "Support"),
        ("Eva",    95, 71000, "North","Sales"),
        ("Frank",  43, 9800,  "South","Support"),
        ("Grace",  78, 33600, "East", "Marketing"),
        ("Henry",  88, 52100, "West", "Sales"),
    ]
    for i, row in enumerate(people):
        bg = LIGHT_PURP if i % 2 == 0 else WHITE
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=5+i, column=ci, value=val)
            c.font = _font(); c.fill = _fill(bg)
            c.alignment = _center(); c.border = _border()

    # Nested IF — Grade
    subhead(ws, 14, "NESTED IF — Performance Grade (Score → A/B/C/D/F)", 6, bg=DARK_TEAL)
    thead(ws, 15, ["Employee","Score","Grade (Nested IF)","Grade (IFS)","Formula (IF)","Formula (IFS)"], bg=DARK_TEAL)
    for i in range(len(people)):
        r = 16 + i
        src = 5 + i
        cell(ws, r, 1, people[i][0], bg=LIGHT_TEAL)
        cell(ws, r, 2, people[i][1], bg=LIGHT_TEAL)
        formula_cell(ws, r, 3, f'=IF(B{src}>=90,"A",IF(B{src}>=80,"B",IF(B{src}>=70,"C",IF(B{src}>=60,"D","F"))))')
        formula_cell(ws, r, 4, f'=IFS(B{src}>=90,"A",B{src}>=80,"B",B{src}>=70,"C",B{src}>=60,"D",TRUE,"F")')
        label_cell(ws, r, 5, '=IF(score>=90,"A",IF(score>=80,"B",IF(score>=70,"C","F")))')
        label_cell(ws, r, 6, '=IFS(score>=90,"A", score>=80,"B", ..., TRUE,"F")')

    # SWITCH — Department label
    subhead(ws, 25, "SWITCH — Map department code to full label", 6, bg=ORANGE)
    thead(ws, 26, ["Employee","Dept","Full Dept Name (SWITCH)","Commission? (AND/OR)","Formula",""], bg=ORANGE)
    for i in range(len(people)):
        r = 27 + i
        src = 5 + i
        cell(ws, r, 1, people[i][0], bg=LIGHT_ORANG)
        cell(ws, r, 2, people[i][3], bg=LIGHT_ORANG)  # region
        formula_cell(ws, r, 3, f'=SWITCH(E{src},"Sales","Sales Team","Marketing","Marketing Dept","Support","Customer Support","Unknown")', bg=LIGHT_ORANG)
        formula_cell(ws, r, 4, f'=IF(AND(C{src}>40000,B{src}>=80),"✅ Eligible","❌ Not Eligible")', bg=LIGHT_ORANG)
        label_cell(ws, r, 5, '=SWITCH(val,"A","Result A","B","Result B","Default")')

    for c, w in [(1,12),(2,10),(3,24),(4,22),(5,42),(6,42)]:
        ws.column_dimensions[get_column_letter(c)].width = w


# ══════════════════════════════════════════════════════════════
#  SHEET 5 — DATE INTELLIGENCE
# ══════════════════════════════════════════════════════════════
def build_dates(ws):
    ws.sheet_view.showGridLines = False
    banner(ws, 1, "📅  Date Intelligence — EOMONTH · NETWORKDAYS · DATEDIF · WEEKDAY · TODAY", 6, bg=DARK_TEAL)

    subhead(ws, 3, "SAMPLE PROJECT DATES", 6, bg=BLUE)
    thead(ws, 4, ["Project","Start Date","End Date","Deadline","Budget","Status"], bg=BLUE)

    projects = [
        ("Dashboard Rebuild",  date(2024,1,15), date(2024,3,20), date(2024,3,31), 15000,"Active"),
        ("Data Pipeline",      date(2024,2,1),  date(2024,5,10), date(2024,4,30), 22000,"Overdue"),
        ("Report Automation",  date(2024,3,10), date(2024,4,25), date(2024,5,15), 8500, "Active"),
        ("ML Model Deploy",    date(2024,4,1),  date(2024,7,30), date(2024,8,1),  45000,"Active"),
        ("CRM Integration",    date(2024,1,5),  date(2024,2,28), date(2024,3,1),  12000,"Complete"),
    ]

    for i, (name, start, end, deadline, budget, status) in enumerate(projects):
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        for ci, val in enumerate([name, start, end, deadline, budget, status], 1):
            c = ws.cell(row=5+i, column=ci, value=val)
            c.font = _font(); c.fill = _fill(bg)
            c.alignment = _center(); c.border = _border()
            if ci in [2, 3, 4]:
                c.number_format = "YYYY-MM-DD"

    # Date formula demos
    subhead(ws, 11, "DATE FORMULA RESULTS", 6, bg=DARK_TEAL)
    thead(ws, 12, ["Project","Duration (days)","Working Days","Month End","Days Overdue","Weekday Name"], bg=DARK_TEAL)

    for i in range(len(projects)):
        r  = 13 + i
        sr = 5  + i
        cell(ws, r, 1, projects[i][0], bg=LIGHT_TEAL)
        formula_cell(ws, r, 2, f'=DATEDIF(B{sr},C{sr},"D")')
        formula_cell(ws, r, 3, f'=NETWORKDAYS(B{sr},C{sr})')
        formula_cell(ws, r, 4, f'=EOMONTH(C{sr},0)', bg=LIGHT_TEAL)
        formula_cell(ws, r, 5, f'=MAX(0,C{sr}-D{sr})')
        formula_cell(ws, r, 6, f'=CHOOSE(WEEKDAY(B{sr},2),"Mon","Tue","Wed","Thu","Fri","Sat","Sun")')
        ws.cell(row=r, column=4).number_format = "YYYY-MM-DD"

    # Formula explanations
    subhead(ws, 19, "FORMULA REFERENCE — DATE FUNCTIONS", 6, bg=BLUE)
    thead(ws, 20, ["Function","Syntax","What It Does","Example","",""], bg=BLUE)
    date_funcs = [
        ("TODAY()",      "=TODAY()",                      "Returns today's date",                              "=TODAY()"),
        ("DATEDIF()",    "=DATEDIF(start,end,'D')",       "Difference between dates in days/months/years",     '=DATEDIF(A1,B1,"D")'),
        ("NETWORKDAYS()","=NETWORKDAYS(start,end)",       "Working days between two dates (excl. weekends)",   "=NETWORKDAYS(A1,B1)"),
        ("EOMONTH()",    "=EOMONTH(date,0)",              "Last day of the month of a given date",             "=EOMONTH(A1,0)"),
        ("WEEKDAY()",    "=WEEKDAY(date,2)",               "Day number of week (2=Mon=1 through Sun=7)",        "=WEEKDAY(A1,2)"),
        ("CHOOSE()",     "=CHOOSE(n,'a','b','c',...)",    "Returns nth item from a list",                      '=CHOOSE(WEEKDAY(A1,2),"Mon","Tue",...)'),
    ]
    for i, row in enumerate(date_funcs):
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        for ci, val in enumerate(row + ("",""), 1):
            c = ws.cell(row=21+i, column=ci, value=val)
            c.font = _font(bold=(ci==1)); c.fill = _fill(bg)
            c.alignment = _left() if ci > 1 else _center()
            c.border = _border()

    for c, w in [(1,22),(2,18),(3,18),(4,16),(5,16),(6,22)]:
        ws.column_dimensions[get_column_letter(c)].width = w


# ══════════════════════════════════════════════════════════════
#  SHEET 6 — DYNAMIC ARRAYS
# ══════════════════════════════════════════════════════════════
def build_dynamic_arrays(ws):
    ws.sheet_view.showGridLines = False
    banner(ws, 1, "📋  Dynamic Arrays — UNIQUE · SORT · FILTER · SEQUENCE", 7, bg=NAVY)

    subhead(ws, 3, "SOURCE DATA (Transactions)", 7, bg=DARK_TEAL)
    thead(ws, 4, ["TxID","Product","Region","Rep","Qty","Revenue","Status"], bg=DARK_TEAL)

    txns = [
        ("T01","Laptop","North","Alice",2,2599.98,"Completed"),
        ("T02","Mouse", "South","Bob",  5,  229.95,"Completed"),
        ("T03","Laptop","North","Alice",1,1299.99,"Returned"),
        ("T04","Chair", "East", "Carol",3,1049.97,"Completed"),
        ("T05","Laptop","West", "David",1,1299.99,"Completed"),
        ("T06","Hub",   "South","Bob",  8,  479.92,"Pending"),
        ("T07","Mouse", "North","Alice",2,   89.98,"Completed"),
        ("T08","Desk",  "East", "Carol",1,  599.99,"Completed"),
        ("T09","Laptop","South","Bob",  2,2599.98,"Completed"),
        ("T10","Chair", "North","Alice",2,  699.98,"Completed"),
        ("T11","Monitor","West","David",3,1199.97,"Completed"),
        ("T12","Hub",   "East", "Carol",5,  299.95,"Returned"),
    ]

    for i, row in enumerate(txns):
        bg = LIGHT_TEAL if i % 2 == 0 else WHITE
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=5+i, column=ci, value=val)
            c.font = _font(); c.fill = _fill(bg)
            c.alignment = _center(); c.border = _border()

    data_end = 5 + len(txns) - 1  # row 16

    # UNIQUE
    subhead(ws, 18, "UNIQUE — Extract unique values from a column", 7, bg=PURPLE)
    cell(ws, 19, 1, "Unique Regions:", bold=True, bg=LIGHT_PURP)
    formula_cell(ws, 19, 2, f'=UNIQUE(C5:C{data_end})', bg=LIGHT_PURP)
    cell(ws, 20, 1, "Unique Products:", bold=True, bg=LIGHT_PURP)
    formula_cell(ws, 20, 2, f'=UNIQUE(B5:B{data_end})', bg=LIGHT_PURP)
    label_cell(ws, 19, 3, '=UNIQUE(range)  →  spills unique values downward automatically')

    # FILTER
    subhead(ws, 22, "FILTER — Return only rows matching a condition", 7, bg=ORANGE)
    cell(ws, 23, 1, "Completed Only:", bold=True, bg=LIGHT_ORANG)
    formula_cell(ws, 23, 2, f'=FILTER(A5:G{data_end},G5:G{data_end}="Completed","No results")', bg=LIGHT_ORANG)
    label_cell(ws, 23, 3, '=FILTER(array, condition, [if_empty])  →  returns matching rows only')

    cell(ws, 25, 1, "Revenue > 500:", bold=True, bg=LIGHT_ORANG)
    formula_cell(ws, 25, 2, f'=FILTER(A5:G{data_end},F5:F{data_end}>500,"No results")', bg=LIGHT_ORANG)

    # SORT & SEQUENCE
    subhead(ws, 27, "SORT & SEQUENCE — Sort data and generate number sequences", 7, bg=BLUE)
    cell(ws, 28, 1, "Sort by Revenue ↓:", bold=True, bg=LIGHT_BLUE)
    formula_cell(ws, 28, 2, f'=SORT(A5:G{data_end},6,-1)', bg=LIGHT_BLUE)
    label_cell(ws, 28, 3, '=SORT(array, sort_col_index, -1 for desc)')

    cell(ws, 30, 1, "Sequence 1–10:", bold=True, bg=LIGHT_BLUE)
    formula_cell(ws, 30, 2, '=SEQUENCE(10)', bg=LIGHT_BLUE)
    cell(ws, 31, 1, "Sequence 1–5 × 1–3 grid:", bold=True, bg=LIGHT_BLUE)
    formula_cell(ws, 31, 2, '=SEQUENCE(5,3,1,1)', bg=LIGHT_BLUE)
    label_cell(ws, 30, 3, '=SEQUENCE(rows, [cols], [start], [step])')

    for c, w in [(1,22),(2,32),(3,52),(4,14),(5,12),(6,14),(7,14)]:
        ws.column_dimensions[get_column_letter(c)].width = w


# ══════════════════════════════════════════════════════════════
#  SHEET 7 — CONDITIONAL FORMATTING
# ══════════════════════════════════════════════════════════════
def build_conditional_formatting(ws):
    ws.sheet_view.showGridLines = False
    banner(ws, 1, "🎨  Conditional Formatting — Colour Scales · Data Bars · Icon Sets · Formula Rules", 6, bg=DARK_TEAL)

    subhead(ws, 3, "SALES PERFORMANCE DATA", 6, bg=DARK_TEAL)
    thead(ws, 4, ["Rep","Q1 Revenue","Q2 Revenue","Q3 Revenue","Q4 Revenue","Annual Total"], bg=DARK_TEAL)

    perf_data = [
        ("Alice",  18500, 22100, 19800, 25600),
        ("Bob",    12300,  9800, 14200, 11500),
        ("Carol",  24100, 26800, 23400, 28900),
        ("David",   8200, 10400,  9100,  7800),
        ("Eva",    29500, 31200, 27800, 33100),
        ("Frank",  15600, 13200, 17400, 16800),
        ("Grace",  21400, 19600, 22800, 24200),
        ("Henry",  11200, 14800, 12600, 13400),
    ]

    for i, (rep, q1, q2, q3, q4) in enumerate(perf_data):
        r  = 5 + i
        bg = LIGHT_TEAL if i % 2 == 0 else WHITE
        cell(ws, r, 1, rep, bg=bg, bold=True)
        for ci, val in enumerate([q1, q2, q3, q4], 2):
            c = ws.cell(row=r, column=ci, value=val)
            c.font = _font(); c.fill = _fill(bg)
            c.alignment = _center(); c.border = _border()
            c.number_format = '"$"#,##0'
        # Annual total formula
        formula_cell(ws, r, 6, f'=SUM(B{r}:E{r})')
        ws.cell(row=r, column=6).number_format = '"$"#,##0'

    # Apply conditional formatting
    # Colour scale on Q1–Q4 (green = high, red = low)
    color_scale = ColorScaleRule(
        start_type="min",  start_color="F1948A",
        mid_type="percentile", mid_value=50, mid_color="FAD7A0",
        end_type="max",    end_color="82E0AA"
    )
    ws.conditional_formatting.add("B5:E12", color_scale)

    # Data bars on Annual Total
    data_bar = DataBarRule(start_type="min", start_value=0,
                           end_type="max",   end_value=None,
                           color="2E86C1")
    ws.conditional_formatting.add("F5:F12", data_bar)

    # Formula rule — highlight top performer row gold
    gold_rule = FormulaRule(
        formula=["$F5=MAX($F$5:$F$12)"],
        fill=_fill("FFF176"),
        font=Font(name="Calibri", bold=True, color="7D6608")
    )
    ws.conditional_formatting.add("A5:F12", gold_rule)

    # Explanations
    subhead(ws, 14, "WHAT EACH FORMATTING RULE DOES", 6, bg=BLUE)
    thead(ws, 15, ["Rule Type","Applied To","Effect","How to Set Up in Excel","",""], bg=BLUE)
    rules = [
        ("3-Colour Scale",    "Q1–Q4 (B:E)", "Red→Yellow→Green based on value",  "Home → Conditional Formatting → Color Scales"),
        ("Data Bars",         "Annual Total (F)", "Blue bar proportional to value",    "Home → Conditional Formatting → Data Bars"),
        ("Formula Rule (Gold)","Entire row",   "Highlights the #1 performer row",  "Home → CF → New Rule → Use formula → =$F5=MAX(...)"),
    ]
    for i, row in enumerate(rules):
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        for ci, val in enumerate(row + ("",""), 1):
            c = ws.cell(row=16+i, column=ci, value=val)
            c.font = _font(bold=(ci==1)); c.fill = _fill(bg)
            c.alignment = _left() if ci > 1 else _center()
            c.border = _border()

    for c, w in [(1,12),(2,14),(3,14),(4,14),(5,14),(6,16)]:
        ws.column_dimensions[get_column_letter(c)].width = w


# ══════════════════════════════════════════════════════════════
#  SHEET 8 — DATA VALIDATION
# ══════════════════════════════════════════════════════════════
def build_data_validation(ws):
    ws.sheet_view.showGridLines = False
    banner(ws, 1, "✅  Data Validation — Dropdown Lists · Numeric Rules · Custom Formulas", 7, bg=GREEN)

    subhead(ws, 3, "INTERACTIVE ORDER ENTRY FORM  (Try editing the yellow cells!)", 7, bg=GREEN)
    cell(ws, 4, 1, "💡 All yellow cells have data validation rules. Try entering invalid values to see error messages.",
         bg=LIGHT_GOLD, bold=False, italic=True, color=GOLD, align="left", border=True, height=28)
    ws.merge_cells("A4:G4")

    # Headers
    thead(ws, 6, ["Order ID","Product","Region","Qty","Unit Price","Discount %","Status"], bg=GREEN)

    # Add 8 editable rows with validation
    sample_orders = [
        ("O001","Laptop Pro 15","North",2,1299.99,0.05,"Completed"),
        ("O002","Webcam HD",    "South",5,  79.99,0.10,"Pending"),
        ("O003","4K Monitor",   "East", 1, 399.99,0.00,"Completed"),
        ("O004","USB-C Hub",    "West", 8,  59.99,0.15,"Returned"),
    ]

    for i, row in enumerate(sample_orders):
        r = 7 + i
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.font = _font(); c.fill = _fill(LIGHT_GOLD)
            c.alignment = _center(); c.border = _border()

    # ── Validation rules ──────────────────────────────────────
    # Product dropdown
    dv_product = DataValidation(
        type="list",
        formula1='"Laptop Pro 15,Wireless Mouse,Mechanical Keyboard,4K Monitor,USB-C Hub,Webcam HD,Ergonomic Chair,Standing Desk"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Product",
        error="Please select a product from the dropdown list."
    )
    ws.add_data_validation(dv_product)
    dv_product.add("B7:B14")

    # Region dropdown
    dv_region = DataValidation(
        type="list",
        formula1='"North,South,East,West,Central"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Region",
        error="Region must be: North, South, East, West, or Central."
    )
    ws.add_data_validation(dv_region)
    dv_region.add("C7:C14")

    # Quantity — whole number 1–100
    dv_qty = DataValidation(
        type="whole",
        operator="between",
        formula1="1", formula2="100",
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Quantity",
        error="Quantity must be a whole number between 1 and 100."
    )
    ws.add_data_validation(dv_qty)
    dv_qty.add("D7:D14")

    # Price — decimal > 0
    dv_price = DataValidation(
        type="decimal",
        operator="greaterThan",
        formula1="0",
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Price",
        error="Unit price must be a positive number."
    )
    ws.add_data_validation(dv_price)
    dv_price.add("E7:E14")

    # Discount — 0 to 0.5
    dv_discount = DataValidation(
        type="decimal",
        operator="between",
        formula1="0", formula2="0.5",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid Discount",
        error="Discount must be between 0 and 0.5 (0% to 50%)."
    )
    ws.add_data_validation(dv_discount)
    dv_discount.add("F7:F14")

    # Status dropdown
    dv_status = DataValidation(
        type="list",
        formula1='"Completed,Pending,Returned,Cancelled"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Status",
        error="Status must be: Completed, Pending, Returned, or Cancelled."
    )
    ws.add_data_validation(dv_status)
    dv_status.add("G7:G14")

    # Validation summary
    subhead(ws, 16, "VALIDATION RULES SUMMARY", 7, bg=DARK_TEAL)
    thead(ws, 17, ["Column","Rule Type","Constraint","Error Message","","",""], bg=DARK_TEAL)
    rules = [
        ("Product (B)",   "List",          "Must be from approved product list",           "Select from dropdown"),
        ("Region (C)",    "List",          "North, South, East, West, Central only",        "Select from dropdown"),
        ("Quantity (D)",  "Whole Number",  "Must be integer between 1 and 100",             "Whole number 1–100"),
        ("Unit Price (E)","Decimal",       "Must be a positive number (> 0)",               "Must be positive"),
        ("Discount (F)",  "Decimal",       "Must be between 0.00 and 0.50",                 "Range: 0–0.5"),
        ("Status (G)",    "List",          "Completed, Pending, Returned, Cancelled only",  "Select from dropdown"),
    ]
    for i, row in enumerate(rules):
        bg = LIGHT_GREEN if i % 2 == 0 else WHITE
        for ci, val in enumerate(row + ("","",""), 1):
            c = ws.cell(row=18+i, column=ci, value=val)
            c.font = _font(bold=(ci==1)); c.fill = _fill(bg)
            c.alignment = _left() if ci > 1 else _center()
            c.border = _border()

    for c, w in [(1,14),(2,22),(3,16),(4,10),(5,14),(6,14),(7,14)]:
        ws.column_dimensions[get_column_letter(c)].width = w


# ══════════════════════════════════════════════════════════════
#  SHEET 9 — FORMULA REFERENCE
# ══════════════════════════════════════════════════════════════
def build_reference(ws):
    ws.sheet_view.showGridLines = False
    banner(ws, 1, "📖  Formula Reference Cheatsheet — All Functions Used in This Workbook", 5, bg=NAVY, size=13)

    ws.merge_cells("A2:E2")
    note = ws["A2"]
    note.value = "Quick reference for every formula demonstrated in this workbook. Copy-paste any syntax directly into Excel."
    note.font  = _font(size=10, italic=True, color=MID_GRAY)
    note.alignment = _center()

    categories = [
        ("LOOKUP & REFERENCE", PURPLE, [
            ("XLOOKUP",       "=XLOOKUP(lookup_value, lookup_array, return_array, [not_found])",
             "Looks up a value in one range and returns the matching value from another. Replaces VLOOKUP."),
            ("INDEX-MATCH",   "=INDEX(return_range, MATCH(lookup, lookup_range, 0))",
             "More flexible than XLOOKUP. Works in any direction. MATCH finds the position, INDEX returns the value."),
            ("IFERROR",       "=IFERROR(formula, value_if_error)",
             "Returns a custom value if a formula produces an error (#N/A, #REF!, etc.). Wrap around any lookup."),
            ("VLOOKUP",       "=VLOOKUP(lookup, table, col_index, FALSE)",
             "Classic lookup — searches leftmost column and returns value from col_index. Use XLOOKUP in modern Excel."),
        ]),
        ("CONDITIONAL AGGREGATION", ORANGE, [
            ("SUMIFS",        "=SUMIFS(sum_range, criteria_range1, criteria1, [range2, criteria2, ...])",
             "Sums values that meet multiple conditions. All ranges must be the same size."),
            ("COUNTIFS",      "=COUNTIFS(criteria_range1, criteria1, [range2, criteria2, ...])",
             "Counts rows that meet multiple conditions simultaneously."),
            ("AVERAGEIFS",    "=AVERAGEIFS(avg_range, criteria_range1, criteria1, ...)",
             "Averages values matching multiple conditions."),
            ("SUMPRODUCT",    "=SUMPRODUCT(array1, array2, ...)",
             "Multiplies corresponding elements and sums the products. Extremely flexible for conditional calculations."),
        ]),
        ("LOGICAL FUNCTIONS", DARK_TEAL, [
            ("IF",            "=IF(condition, value_if_true, value_if_false)",
             "Basic conditional. Nest multiple IFs for tiered logic (avoid more than 3 levels — use IFS instead)."),
            ("IFS",           "=IFS(cond1, val1, cond2, val2, ..., TRUE, default)",
             "Tests multiple conditions in sequence. Cleaner than nested IF. The TRUE at end provides the default."),
            ("SWITCH",        "=SWITCH(expression, val1, result1, val2, result2, ..., default)",
             "Matches an expression to a list of values and returns the corresponding result."),
            ("AND / OR",      "=AND(cond1, cond2, ...) / =OR(cond1, cond2, ...)",
             "AND returns TRUE only if ALL conditions are true. OR returns TRUE if ANY condition is true."),
        ]),
        ("DATE FUNCTIONS", BLUE, [
            ("TODAY / NOW",   "=TODAY() / =NOW()",
             "Returns today's date or current date-time. Recalculates every time the sheet refreshes."),
            ("DATEDIF",       '=DATEDIF(start_date, end_date, "D"/"M"/"Y")',
             'Calculates the difference between dates. "D"=days, "M"=months, "Y"=years. Note: hidden function (not in autocomplete).'),
            ("NETWORKDAYS",   "=NETWORKDAYS(start_date, end_date, [holidays])",
             "Returns the number of working days (Mon–Fri) between two dates, excluding weekends and optional holidays."),
            ("EOMONTH",       "=EOMONTH(start_date, months)",
             "Returns the last day of the month. EOMONTH(date,0)=this month end. EOMONTH(date,1)=next month end."),
            ("WEEKDAY",       "=WEEKDAY(date, return_type)",
             "Returns a number for the day of week. Use return_type=2 for Mon=1 through Sun=7."),
        ]),
        ("DYNAMIC ARRAYS (Excel 365)", NAVY, [
            ("UNIQUE",        "=UNIQUE(array, [by_col], [exactly_once])",
             "Returns a list of unique values from a range. Result spills automatically into adjacent cells."),
            ("FILTER",        "=FILTER(array, include, [if_empty])",
             "Returns only the rows/columns that match a condition. The include argument is a TRUE/FALSE array."),
            ("SORT",          "=SORT(array, [sort_index], [sort_order], [by_col])",
             "Returns a sorted version of an array. sort_order: 1=ascending, -1=descending."),
            ("SEQUENCE",      "=SEQUENCE(rows, [cols], [start], [step])",
             "Generates a sequential number array. =SEQUENCE(10) gives 1 to 10 in a column."),
        ]),
    ]

    row_idx = 4
    for cat_name, cat_color, funcs in categories:
        subhead(ws, row_idx, cat_name, 5, bg=cat_color)
        row_idx += 1
        thead(ws, row_idx, ["Function","Syntax","Plain-English Explanation","",""], bg=cat_color)
        row_idx += 1
        for j, (fname, syntax, explanation) in enumerate(funcs):
            bg = LIGHT_TEAL if j % 2 == 0 else WHITE
            cell(ws, row_idx, 1, fname,       bg=bg, bold=True,   color=cat_color)
            cell(ws, row_idx, 2, syntax,      bg=bg, italic=True, align="left")
            ws.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=5)
            c = ws.cell(row=row_idx, column=3, value=explanation)
            c.font = _font(size=10, color=DARK_GRAY); c.fill = _fill(bg)
            c.alignment = _left(); c.border = _border()
            ws.row_dimensions[row_idx].height = 22
            row_idx += 1
        row_idx += 1

    for c, w in [(1,22),(2,55),(3,55),(4,10),(5,10)]:
        ws.column_dimensions[get_column_letter(c)].width = w


# ══════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════
def build():
    print("\n" + "═"*58)
    print("   EXCEL ADVANCED FUNCTIONS SHOWCASE BUILDER")
    print("═"*58)

    wb = Workbook()
    wb.remove(wb.active)

    sheets = [
        ("🏠 Overview",          build_overview),
        ("📦 Product Lookup",     build_lookup),
        ("📊 Sales Summary",      build_sales_summary),
        ("🔢 Nested Logic",       build_nested_logic),
        ("📅 Date Intelligence",  build_dates),
        ("📋 Dynamic Arrays",     build_dynamic_arrays),
        ("🎨 Cond. Formatting",   build_conditional_formatting),
        ("✅ Data Validation",    build_data_validation),
        ("📖 Formula Reference",  build_reference),
    ]

    for name, builder in sheets:
        print(f"🔨  Building: {name}")
        ws = wb.create_sheet(title=name)
        ws.sheet_view.showGridLines = False
        builder(ws)

    # Tab colours
    tab_colors = ["1ABC9C","E74C3C","E67E22","8E44AD","2E86C1","1A5276","D35400","27AE60","2C3E50"]
    for ws, color in zip(wb.worksheets, tab_colors):
        ws.sheet_properties.tabColor = color

    out = "Excel_Advanced_Functions_Showcase.xlsx"
    wb.save(out)
    size_kb = round(os.path.getsize(out)/1024, 1)
    print(f"\n✅  Saved → {out}  ({size_kb} KB)")
    print("═"*58 + "\n")

if __name__ == "__main__":
    build()
